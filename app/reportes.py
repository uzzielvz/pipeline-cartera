from flask import Blueprint, render_template, request, send_file, flash, redirect, url_for, Response
from app.models import db, ReportHistory
from flask_login import current_user, login_required
from app.auth import require_permission
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
import os
import re
import logging
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
import urllib.parse
from config import (
    ALLOWED_EXTENSIONS, UPLOAD_FOLDER, REPORTS_FOLDER, MAX_FILE_SIZE, COLUMN_MAPPING, 
    DTYPE_CONFIG, LISTA_FRAUDE, CODIGOS_RECUPERADOR_EXCLUIR, PERIODICIDAD_A_DIAS, EXCEL_CONFIG, COLORS, ADDITIONAL_COLUMNS,
    MORA_BLUE_COLUMNS, CURRENCY_COLUMNS_KEYWORDS, DATE_COLUMNS_KEYWORDS
)

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

reportes_bp = Blueprint('reportes', __name__)

def allowed_file(filename):
    """Verifica que el archivo sea Excel"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def move_to_reports_folder(file_path, report_type='individual'):
    """Mueve un archivo generado al directorio de reportes permanentes"""
    try:
        # Crear directorio si no existe
        os.makedirs(REPORTS_FOLDER, exist_ok=True)
        
        # Generar nombre único para el archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = os.path.basename(file_path)
        name, ext = os.path.splitext(filename)
        
        # Crear nuevo nombre con timestamp y tipo
        new_filename = f"{name}_{report_type}_{timestamp}{ext}"
        new_path = os.path.join(REPORTS_FOLDER, new_filename)
        
        # Mover archivo
        import shutil
        shutil.move(file_path, new_path)
        
        logger.info(f"✅ Archivo movido a directorio de reportes: {new_path}")
        return new_path
        
    except Exception as e:
        logger.error(f"❌ Error moviendo archivo a directorio de reportes: {str(e)}")
        return file_path  # Retornar ruta original si hay error

def validate_file_size(file_path):
    """Valida que el archivo no exceda el tamaño máximo permitido"""
    try:
        file_size = os.path.getsize(file_path)
        if file_size > MAX_FILE_SIZE:
            raise ValueError(f"El archivo excede el tamaño máximo de {MAX_FILE_SIZE // (1024*1024)}MB")
        return True
    except OSError as e:
        raise ValueError(f"Error al verificar el tamaño del archivo: {str(e)}")

def clean_dataframe_columns(df):
    """Limpia los nombres de columnas del DataFrame"""
    df.columns = df.columns.str.replace('\n', ' ').str.strip()
    return df

def standardize_codes(df, code_columns):
    """Estandariza códigos a 6 dígitos con ceros a la izquierda"""
    for col in code_columns:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int).astype(str).str.zfill(6)
    return df

def clean_phone_numbers(df):
    """Limpia y estandariza números de teléfono"""
    df = df.copy()  # Crear copia para evitar SettingWithCopyWarning
    for col in df.columns:
        if 'Teléfono' in col:
            df[col] = df[col].fillna('').astype(str)
    return df

def add_geolocation_links(df, geolocation_column):
    """Añade columnas de enlaces de geolocalización al DataFrame"""
    df = df.copy()  # Crear copia para evitar SettingWithCopyWarning
    if geolocation_column in df.columns:
        geolocalizacion_data = df[geolocation_column].apply(generar_link_google_maps)
        df['link_texto'] = [item[0] for item in geolocalizacion_data]
        df['link_url'] = [item[1] for item in geolocalizacion_data]
    return df


def add_par_column(df, mora_column):
    """Añade columna PAR y la posiciona correctamente"""
    df = df.copy()  # Crear copia para evitar SettingWithCopyWarning
    
    # Solo eliminar columnas que sean exactamente "PAR" o variantes exactas de "PAR 2"
    columnas_par_exactas = []
    for col in df.columns:
        col_str = str(col).strip().lower()
        # Solo eliminar si es exactamente "par" o variantes de "par 2"
        if col_str == 'par' or col_str in ['par 2', 'par2', 'par  2', 'par   2', 'par-2', 'par_2', 'par.2']:
            columnas_par_exactas.append(col)
    
    if columnas_par_exactas:
        logger.warning(f"⚠️ ELIMINANDO columnas PAR exactas: {columnas_par_exactas}")
        df = df.drop(columns=columnas_par_exactas, errors='ignore')
    
    # Crear la columna PAR
    df['PAR'] = df[mora_column].apply(asignar_rango_mora)
    logger.info(f"✅ PAR creado")
    
    # Reordenar columnas para que 'PAR' esté al lado de 'Días de mora'
    columnas = df.columns.tolist()
    mora_index = columnas.index(mora_column)
    nuevas_columnas = (columnas[:mora_index+1] + 
                      ['PAR'] + 
                      columnas[mora_index+1:])
    
    return df[nuevas_columnas]


def generate_valid_table_name(sheet_name):
    """
    Genera un nombre de tabla válido para Excel que siempre empiece con una letra.
    
    Reglas de Excel para nombres de tabla:
    - Debe empezar con una letra o guión bajo
    - No puede empezar con un número
    - Solo puede contener letras, números, guiones bajos y puntos
    - No puede contener espacios
    - Máximo 255 caracteres
    """
    # Limpiar el nombre de la hoja
    clean_name = str(sheet_name).replace(' ', '_').replace('-', '_')
    
    # Remover caracteres no válidos (mantener solo letras, números, guiones bajos y puntos)
    import re
    clean_name = re.sub(r'[^a-zA-Z0-9_.]', '', clean_name)
    
    # Asegurar que empiece con una letra o guión bajo
    if clean_name and clean_name[0].isdigit():
        clean_name = f"T_{clean_name}"
    elif not clean_name or not clean_name[0].isalpha():
        clean_name = f"T_{clean_name}" if clean_name else "T_Table"
    
    # Limitar longitud a 255 caracteres
    if len(clean_name) > 255:
        clean_name = clean_name[:255]
    
    return clean_name


def generar_link_google_maps(geolocalizacion):
    """
    Traductor de Direcciones - Convierte cualquier formato de geolocalización en enlaces de Google Maps.
    
    Casos manejados:
    - URL existente de Google Maps: ("Ver en mapa", url_original)
    - Coordenadas GPS: ("Ver en mapa", url_búsqueda)
    - Dirección de texto: ("Ver en mapa", url_búsqueda)
    - Vacío/nulo: ("Ver en mapa", url_generico)
    """
    # Caso 1: Valor vacío o nulo
    if pd.isna(geolocalizacion) or str(geolocalizacion).strip() == '':
        return ("Ver en mapa", "https://maps.google.com")
    
    geolocalizacion = str(geolocalizacion).strip()
    
    # Caso 2: Ya es una URL de Google Maps
    if "http" in geolocalizacion or "google.com/maps" in geolocalizacion:
        return ("Ver en mapa", geolocalizacion)
    
    # Caso 3: Contiene coordenadas GPS (patrón con °, ', ", N, W, S, E)
    if any(char in geolocalizacion for char in ['°', "'", '"', 'N', 'W', 'S', 'E']):
        try:
            # Extraer coordenadas usando regex
            # Patrón para coordenadas: 19°12'12.2"N 100°07'51.8"W
            coord_pattern = r"(\d+)°(\d+)'([\d.]+)\"([NS])\s+(\d+)°(\d+)'([\d.]+)\"([WE])"
            match = re.search(coord_pattern, geolocalizacion)
            
            if match:
                lat_deg, lat_min, lat_sec, lat_dir = match.groups()[:4]
                lon_deg, lon_min, lon_sec, lon_dir = match.groups()[4:]
                
                # Convertir a decimal
                lat_decimal = float(lat_deg) + float(lat_min)/60 + float(lat_sec)/3600
                lon_decimal = float(lon_deg) + float(lon_min)/60 + float(lon_sec)/3600
                
                # Aplicar dirección (N/S, E/W)
                if lat_dir == 'S':
                    lat_decimal = -lat_decimal
                if lon_dir == 'W':
                    lon_decimal = -lon_decimal
                
                # Crear URL de Google Maps
                url = f"https://www.google.com/maps/search/?api=1&query={lat_decimal},{lon_decimal}"
                return ("Ver en mapa", url)
        except:
            pass
    
    # Caso 4: Dirección de texto - crear búsqueda
    direccion_encoded = urllib.parse.quote_plus(geolocalizacion)
    url = f"https://www.google.com/maps/search/?api=1&query={direccion_encoded}"
    return ("Ver en mapa", url)

def asignar_rango_mora(dias_mora):
    """
    Asigna valor PAR (Período de Antigüedad de Recuperación) basado en días de mora.

    Reglas de categorización:
    - 0 días o sin mora: '0'
    - 1-7 días: '7'
    - 8-15 días: '15'
    - 16-30 días: '30'
    - 31-60 días: '60'
    - 61-90 días: '90'
    - 91-180 días: 'Mayor_90'
    - >180 días: 'Mayor_180'
    """
    if pd.isna(dias_mora) or dias_mora < 1:
        return '0'
    elif dias_mora <= 7:
        return '7'
    elif dias_mora <= 15:
        return '15'
    elif dias_mora <= 30:
        return '30'
    elif dias_mora <= 60:
        return '60'
    elif dias_mora <= 90:
        return '90'
    elif dias_mora <= 180:
        return 'Mayor_90'
    else:
        return 'Mayor_180'

def escribir_hipervinculo_excel(worksheet, row, col, texto, url):
    """Escribe un hipervínculo en una celda de Excel usando fórmula HYPERLINK (más confiable con openpyxl)."""
    cell = worksheet.cell(row=row, column=col)
    if url and pd.notna(url) and str(url).strip():
        # Escapar comillas dobles en texto y url para la fórmula
        url_safe = str(url).replace('"', '""')
        texto_safe = str(texto).replace('"', '""') if pd.notna(texto) and str(texto).strip() else 'Link'
        cell.value = f'=HYPERLINK("{url_safe}","{texto_safe}")'
        cell.font = Font(color="0000FF", underline="single")

def generar_concepto_deposito(df):
    """
    Genera la columna 'Concepto Depósito' con formato: 1 + código_acreditado(6 dígitos) + ciclo(2 dígitos)
    Ejemplo: 100100401 (1 + 001004 + 01)
    """
    if 'Código acreditado' not in df.columns or 'Ciclo' not in df.columns:
        logger.warning("⚠️ No se puede generar 'Concepto Depósito': faltan columnas 'Código acreditado' o 'Ciclo'")
        return pd.Series([''] * len(df))
    
    # Asegurar que Código acreditado tenga 6 dígitos y Ciclo tenga 2 dígitos
    codigo = df['Código acreditado'].astype(str).str.strip().str.zfill(6)
    ciclo = df['Ciclo'].astype(str).str.strip().str.zfill(2)
    
    # Generar concepto: 1 + código(6) + ciclo(2) = 9 dígitos total
    concepto_deposito = '1' + codigo + ciclo
    
    return concepto_deposito

def agregar_columna_concepto_deposito(df):
    """
    Agrega la columna 'Concepto Depósito' después de 'Forma de entrega' si existe,
    o al final si no existe.
    Solo asigna valor al registro con el ciclo mayor cuando hay duplicados del mismo código.
    """
    # Verificar que existan las columnas necesarias
    if 'Código acreditado' not in df.columns or 'Ciclo' not in df.columns:
        logger.warning("⚠️ No se puede generar 'Concepto Depósito': faltan columnas 'Código acreditado' o 'Ciclo'")
        concepto = pd.Series([''] * len(df))
    else:
        # Asegurar que Código acreditado tenga 6 dígitos y Ciclo tenga 2 dígitos
        codigo = df['Código acreditado'].astype(str).str.strip().str.zfill(6)
        ciclo_str = df['Ciclo'].astype(str).str.strip().str.zfill(2)
        ciclo_num = pd.to_numeric(df['Ciclo'], errors='coerce').fillna(0)
        
        # Generar concepto temporalmente
        concepto_temporal = '1' + codigo + ciclo_str
        
        # Para cada código, identificar la fila con el ciclo mayor
        concepto = pd.Series([''] * len(df))
        
        for codigo_unico in codigo.unique():
            # Encontrar todas las filas con este código
            mascara = codigo == codigo_unico
            indices = df.index[mascara]
            
            if len(indices) > 1:
                # Hay duplicados: encontrar el índice con el ciclo mayor
                ciclos_valores = ciclo_num.loc[indices]
                indice_ciclo_mayor = ciclos_valores.idxmax()
                concepto.loc[indice_ciclo_mayor] = concepto_temporal.loc[indice_ciclo_mayor]
                logger.info(f"🔍 Código {codigo_unico}: {len(indices)} duplicados, asignado a ciclo {ciclo_str.loc[indice_ciclo_mayor]}")
            else:
                # No hay duplicados: asignar normalmente
                concepto.loc[indices[0]] = concepto_temporal.loc[indices[0]]
    
    # Buscar la columna 'Forma de entrega'
    columna_forma_entrega = None
    for col in df.columns:
        if 'forma de entrega' in str(col).lower() or 'forma entrega' in str(col).lower():
            columna_forma_entrega = col
            break
    
    if columna_forma_entrega and columna_forma_entrega in df.columns:
        # Insertar después de 'Forma de entrega'
        forma_index = df.columns.get_loc(columna_forma_entrega)
        df.insert(forma_index + 1, 'Concepto Depósito', concepto)
        logger.info(f"✅ Columna 'Concepto Depósito' agregada después de '{columna_forma_entrega}'")
    else:
        # Agregar al final si no se encuentra 'Forma de entrega'
        df['Concepto Depósito'] = concepto
        logger.info("✅ Columna 'Concepto Depósito' agregada al final (no se encontró 'Forma de entrega')")
    
    return df

def agregar_columnas_riesgo_y_mora(df):
    """
    Agrega las columnas 'Saldo riesgo capital', 'Saldo riesgo total' y '% MORA' 
    después de 'Concepto Depósito' si existe, o al final si no existe.
    
    Fórmulas:
    - Saldo riesgo capital = IF(Días de mora > 0, Saldo capital, 0)
    - Saldo riesgo total = IF(Días de mora > 0, Saldo total, 0)
    - % MORA = Saldo vencido / Saldo total
    """
    # Verificar que existan las columnas necesarias
    columnas_requeridas = ['Días de mora', 'Saldo capital', 'Saldo total', 'Saldo vencido']
    columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
    
    if columnas_faltantes:
        logger.warning(f"⚠️ No se pueden generar columnas de riesgo: faltan columnas {columnas_faltantes}")
        # Crear columnas vacías
        df['Saldo riesgo capital'] = 0
        df['Saldo riesgo total'] = 0
        df['% MORA'] = 0
        return df
    
    # Calcular Saldo riesgo capital = IF(Días de mora > 0, Saldo capital, 0)
    saldo_riesgo_capital = df.apply(
        lambda row: row['Saldo capital'] if pd.notna(row['Días de mora']) and row['Días de mora'] > 0 else 0,
        axis=1
    )
    
    # Calcular Saldo riesgo total = IF(Días de mora > 0, Saldo total, 0)
    saldo_riesgo_total = df.apply(
        lambda row: row['Saldo total'] if pd.notna(row['Días de mora']) and row['Días de mora'] > 0 else 0,
        axis=1
    )
    
    # Calcular % MORA = Saldo vencido / Saldo total
    pct_mora = df.apply(
        lambda row: (row['Saldo vencido'] / row['Saldo total']) 
                    if pd.notna(row['Saldo total']) and row['Saldo total'] != 0 
                    else 0,
        axis=1
    )
    
    # Buscar la columna 'Concepto Depósito' para insertar después
    if 'Concepto Depósito' in df.columns:
        concepto_index = df.columns.get_loc('Concepto Depósito')
        # Insertar las 3 columnas después de 'Concepto Depósito'
        df.insert(concepto_index + 1, 'Saldo riesgo capital', saldo_riesgo_capital)
        df.insert(concepto_index + 2, 'Saldo riesgo total', saldo_riesgo_total)
        df.insert(concepto_index + 3, '% MORA', pct_mora)
        logger.info("✅ Columnas 'Saldo riesgo capital', 'Saldo riesgo total' y '% MORA' agregadas después de 'Concepto Depósito'")
    else:
        # Agregar al final si no se encuentra 'Concepto Depósito'
        df['Saldo riesgo capital'] = saldo_riesgo_capital
        df['Saldo riesgo total'] = saldo_riesgo_total
        df['% MORA'] = pct_mora
        logger.info("✅ Columnas 'Saldo riesgo capital', 'Saldo riesgo total' y '% MORA' agregadas al final")
    
    return df

def _normalizar_texto_para_mapeo(s):
    """Normaliza texto para búsqueda en mapeo (minúsculas, sin tildes)."""
    if pd.isna(s):
        return ''
    s = str(s).strip().lower()
    import unicodedata
    s = unicodedata.normalize('NFD', s)
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn')

def agregar_columnas_dias_ultimo_pago_y_alerta(df):
    """
    Agrega columnas 'Días desde el último pago' y 'Alerta' al DataFrame.
    Días = hoy - Último pago. Alerta = Sí si días > plazo (según Periodicidad).
    Inserta después de '% MORA' si existe.
    """
    col_ultimo_pago = COLUMN_MAPPING.get('ultimo_pago', 'Último pago')
    col_periodicidad = COLUMN_MAPPING.get('periodicidad', 'Periodicidad')
    hoy = pd.Timestamp.now().normalize()

    # 1. Días desde el último pago
    if col_ultimo_pago in df.columns:
        fechas = pd.to_datetime(df[col_ultimo_pago], errors='coerce')
        dias_desde_ultimo = (hoy - fechas).dt.days
    else:
        dias_desde_ultimo = pd.Series([None] * len(df))

    # 2. Plazo en días por fila (Periodicidad -> días)
    def periodo_a_dias(val):
        if pd.isna(val):
            return 30
        try:
            v_num = float(val)
            if v_num > 0:
                return int(v_num)
        except (ValueError, TypeError):
            pass
        k = _normalizar_texto_para_mapeo(val)
        return PERIODICIDAD_A_DIAS.get(k, 30)

    if col_periodicidad in df.columns:
        plazo_dias = df[col_periodicidad].apply(periodo_a_dias)
    else:
        plazo_dias = pd.Series([30] * len(df))

    # 3. Alerta = 1 si días_desde_ultimo > plazo_dias, 0 en caso contrario
    alerta = pd.Series([0] * len(df), index=df.index)
    mask = dias_desde_ultimo.notna() & (dias_desde_ultimo > plazo_dias)
    alerta.loc[mask] = 1

    # 4. Insertar después de '% MORA'
    df = df.copy()
    df['Días desde el último pago'] = dias_desde_ultimo
    df['Alerta'] = alerta
    if '% MORA' in df.columns:
        cols = df.columns.tolist()
        cols.remove('Días desde el último pago')
        cols.remove('Alerta')
        idx = cols.index('% MORA') + 1
        df = df[cols[:idx] + ['Días desde el último pago', 'Alerta'] + cols[idx:]]
    logger.info("✅ Columnas 'Días desde el último pago' y 'Alerta' agregadas")
    return df

def agregar_columnas_nuevas(df):
    """
    Agrega las 3 columnas nuevas al final del DataFrame (iteración 3):
    - Cuotas sin pagar  = Días desde el último pago / días de periodicidad (decimal)
    - Saldo_Riesgo_total = Saldo total si mora > 30, sino 0  (nueva definición)
    - Combinado         = cuotas sin pagar redondeadas si mora <= 30, sino Saldo_Riesgo_total
    """
    col_mora         = 'Días de mora'
    col_dias_pago    = 'Días desde el último pago'
    col_periodicidad = COLUMN_MAPPING.get('periodicidad', 'Periodicidad')
    col_saldo_total  = 'Saldo total'
    df = df.copy()

    def _periodo_a_dias(val):
        if pd.isna(val):
            return 30
        try:
            v = float(val)
            if v > 0:
                return int(v)
        except (ValueError, TypeError):
            pass
        return PERIODICIDAD_A_DIAS.get(_normalizar_texto_para_mapeo(val), 30)

    plazo = df[col_periodicidad].apply(_periodo_a_dias) if col_periodicidad in df.columns else pd.Series([30]*len(df), index=df.index)
    dias_pago = df[col_dias_pago] if col_dias_pago in df.columns else pd.Series([0]*len(df), index=df.index)
    mora      = df[col_mora]      if col_mora      in df.columns else pd.Series([0]*len(df), index=df.index)
    saldo_t   = df[col_saldo_total] if col_saldo_total in df.columns else pd.Series([0]*len(df), index=df.index)

    cuotas = dias_pago / plazo.replace(0, 1)

    saldo_riesgo_nuevo = saldo_t.where(mora > 30, other=0)

    combinado = cuotas.round(0).where(mora <= 30, other=saldo_riesgo_nuevo)

    df['Cuotas sin pagar']   = cuotas
    df['Saldo_Riesgo_total'] = saldo_riesgo_nuevo
    df['Combinado']          = combinado

    logger.info("✅ Columnas 'Cuotas sin pagar', 'Saldo_Riesgo_total' y 'Combinado' agregadas")
    return df


def limpiar_celda_segura(cell):
    """
    Limpia una celda de forma segura, verificando si es MergedCell (solo lectura).
    
    Args:
        cell: Objeto Cell de openpyxl
        
    Returns:
        True si se pudo limpiar, False si es MergedCell
    """
    try:
        from openpyxl.cell.cell import MergedCell
        if isinstance(cell, MergedCell):
            return False  # No se puede modificar MergedCell
        cell.value = None
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell.border = Border()
        return True
    except (AttributeError, TypeError, ImportError):
        # Si hay error, intentar solo limpiar el valor
        try:
            if not isinstance(cell, MergedCell):
                cell.value = None
                return True
        except:
            pass
        return False

def crear_hoja_x_coordinacion(df_completo):
    """
    Crea la hoja 'X_Coordinación' con datos agregados por coordinación.
    
    Estructura:
    - Filas 1-4: Vacías
    - Fila 5: Solo "PAR" en columna 11
    - Fila 6: Encabezados principales (Coordinación, Cantidad Prestada, etc.)
    - Fila 7: Encabezados de rangos de días (Etiquetas de fila, 0, 7, 15, 30, 60, 90, Mayor_90, etc.)
    - Fila 8+: Datos por coordinación
    - Última fila: Total general
    
    Args:
        df_completo: DataFrame completo con todos los datos procesados
        
    Returns:
        DataFrame con la estructura de la hoja X_Coordinación
    """
    logger.info("Creando hoja X_Coordinación...")
    
    # Verificar columnas requeridas
    columna_coordinacion = COLUMN_MAPPING.get('coordinacion', 'Coordinación')
    columnas_requeridas = [
        columna_coordinacion,
        'Cantidad Prestada',
        'Saldo capital',
        'Saldo vencido',
        'Saldo total',
        'Saldo riesgo capital',
        'Saldo riesgo total',
        '% MORA',
        'Días de mora'
    ]
    
    # Verificar y calcular columnas de riesgo si no existen
    if 'Saldo riesgo capital' not in df_completo.columns:
        logger.info("🔍 Calculando 'Saldo riesgo capital' (no existe en df_completo)")
        df_completo['Saldo riesgo capital'] = df_completo.apply(
            lambda row: row['Saldo capital'] if pd.notna(row.get('Días de mora', 0)) and row.get('Días de mora', 0) > 0 else 0,
            axis=1
        )
    
    if 'Saldo riesgo total' not in df_completo.columns:
        logger.info("🔍 Calculando 'Saldo riesgo total' (no existe en df_completo)")
        df_completo['Saldo riesgo total'] = df_completo.apply(
            lambda row: row['Saldo total'] if pd.notna(row.get('Días de mora', 0)) and row.get('Días de mora', 0) > 0 else 0,
            axis=1
        )
    
    if '% MORA' not in df_completo.columns:
        logger.info("🔍 Calculando '% MORA' (no existe en df_completo)")
        df_completo['% MORA'] = df_completo.apply(
            lambda row: (row['Saldo vencido'] / row['Saldo total']) 
                        if pd.notna(row.get('Saldo total', 0)) and row.get('Saldo total', 0) != 0 
                        else 0,
            axis=1
        )
    
    # Verificar columnas requeridas después de calcular las que faltaban
    columnas_faltantes = [col for col in columnas_requeridas if col not in df_completo.columns]
    if columnas_faltantes:
        logger.warning(f"⚠️ Columnas faltantes para X_Coordinación después de calcular: {columnas_faltantes}")
        logger.warning(f"🔍 Columnas disponibles en df_completo: {list(df_completo.columns)[:30]}")
        logger.warning(f"🔍 Total columnas: {len(df_completo.columns)}")
        # Crear DataFrame vacío con estructura mínima
        return pd.DataFrame()
    
    logger.info(f"✅ Todas las columnas requeridas están presentes. Total registros: {len(df_completo)}")
    
    # Agrupar por Coordinación y calcular agregaciones
    grupo = df_completo.groupby(columna_coordinacion, dropna=False).agg({
        'Cantidad Prestada': 'sum',
        'Saldo capital': 'sum',
        'Saldo vencido': 'sum',
        'Saldo total': 'sum',
        'Saldo riesgo capital': 'sum',
        'Saldo riesgo total': 'sum',
        '% MORA': 'mean'  # Promedio del porcentaje (se recalculará después)
    }).reset_index()
    
    # Calcular % MORA correcto (promedio ponderado o recálculo)
    # % MORA = Saldo vencido / Saldo total (por coordinación)
    grupo['% MORA'] = grupo.apply(
        lambda row: (row['Saldo vencido'] / row['Saldo total']) 
                    if pd.notna(row['Saldo total']) and row['Saldo total'] != 0 
                    else 0,
        axis=1
    )
    
    # Calcular rangos de días de mora para cada coordinación
    def calcular_rangos_mora(df_coord, columna_mora='Días de mora', columna_riesgo='Saldo riesgo total'):
        """Calcula la suma de Saldo riesgo total por rangos de días de mora"""
        rangos = {
            '0': 0,
            '1-7': 0,
            '8-15': 0,
            '16-30': 0,
            '31-60': 0,
            '61-90': 0,
            'Mayor_90': 0
        }
        
        if columna_mora not in df_coord.columns or columna_riesgo not in df_coord.columns:
            return rangos
        
        for idx, row in df_coord.iterrows():
            dias_mora = row[columna_mora] if pd.notna(row[columna_mora]) else 0
            saldo_riesgo = row[columna_riesgo] if pd.notna(row[columna_riesgo]) else 0
            
            if dias_mora == 0:
                rangos['0'] += saldo_riesgo
            elif 1 <= dias_mora <= 7:
                rangos['1-7'] += saldo_riesgo
            elif 8 <= dias_mora <= 15:
                rangos['8-15'] += saldo_riesgo
            elif 16 <= dias_mora <= 30:
                rangos['16-30'] += saldo_riesgo
            elif 31 <= dias_mora <= 60:
                rangos['31-60'] += saldo_riesgo
            elif 61 <= dias_mora <= 90:
                rangos['61-90'] += saldo_riesgo
            else:
                rangos['Mayor_90'] += saldo_riesgo
        
        return rangos
    
    # Calcular rangos para cada coordinación
    rangos_por_coord = []
    for coord in grupo[columna_coordinacion]:
        df_coord = df_completo[df_completo[columna_coordinacion] == coord]
        rangos = calcular_rangos_mora(df_coord)
        rangos_por_coord.append(rangos)
    
    # Agregar columnas de rangos al grupo
    grupo['Rango_0'] = [r['0'] for r in rangos_por_coord]
    grupo['Rango_1-7'] = [r['1-7'] for r in rangos_por_coord]
    grupo['Rango_8-15'] = [r['8-15'] for r in rangos_por_coord]
    grupo['Rango_16-30'] = [r['16-30'] for r in rangos_por_coord]
    grupo['Rango_31-60'] = [r['31-60'] for r in rangos_por_coord]
    grupo['Rango_61-90'] = [r['61-90'] for r in rangos_por_coord]
    grupo['Rango_Mayor_90'] = [r['Mayor_90'] for r in rangos_por_coord]
    
    # Calcular total general
    total_general = {
        columna_coordinacion: 'Total',
        'Cantidad Prestada': grupo['Cantidad Prestada'].sum(),
        'Saldo capital': grupo['Saldo capital'].sum(),
        'Saldo vencido': grupo['Saldo vencido'].sum(),
        'Saldo total': grupo['Saldo total'].sum(),
        'Saldo riesgo capital': grupo['Saldo riesgo capital'].sum(),
        'Saldo riesgo total': grupo['Saldo riesgo total'].sum(),
        '% MORA': (grupo['Saldo vencido'].sum() / grupo['Saldo total'].sum()) 
                  if grupo['Saldo total'].sum() != 0 else 0,
        'Rango_0': grupo['Rango_0'].sum(),
        'Rango_1-7': grupo['Rango_1-7'].sum(),
        'Rango_8-15': grupo['Rango_8-15'].sum(),
        'Rango_16-30': grupo['Rango_16-30'].sum(),
        'Rango_31-60': grupo['Rango_31-60'].sum(),
        'Rango_61-90': grupo['Rango_61-90'].sum(),
        'Rango_Mayor_90': grupo['Rango_Mayor_90'].sum()
    }
    
    # Crear DataFrame final con estructura específica
    # Necesitamos crear un DataFrame que se escriba empezando en fila 8
    # Las filas 1-7 se escribirán manualmente en Excel
    
    # Preparar datos para escribir (sin las filas vacías iniciales)
    df_resultado = grupo.copy()
    
    # Renombrar columnas para que coincidan con el formato esperado
    df_resultado = df_resultado.rename(columns={
        columna_coordinacion: 'Coordinación',
        'Rango_0': 'Rango_0',
        'Rango_1-7': 'Rango_1-7',
        'Rango_8-15': 'Rango_8-15',
        'Rango_16-30': 'Rango_16-30',
        'Rango_31-60': 'Rango_31-60',
        'Rango_61-90': 'Rango_61-90',
        'Rango_Mayor_90': 'Rango_Mayor_90'
    })
    
    # Agregar fila de total general
    fila_total = pd.DataFrame([total_general])
    df_resultado = pd.concat([df_resultado, fila_total], ignore_index=True)
    
    logger.info(f"✅ Hoja X_Coordinación creada con {len(grupo)} coordinaciones + 1 total")
    
    return df_resultado

def crear_hoja_x_recuperador(df_completo):
    """
    Crea la hoja 'X_Recuperador' con datos agregados por coordinación y recuperador.
    
    Similar a X_Coordinación pero agrupando por Coordinación + Recuperador.
    
    Args:
        df_completo: DataFrame completo con todos los datos procesados
        
    Returns:
        DataFrame con la estructura de la hoja X_Recuperador
    """
    logger.info("Creando hoja X_Recuperador...")
    
    # Verificar columnas requeridas
    columna_coordinacion = COLUMN_MAPPING.get('coordinacion', 'Coordinación')
    columnas_requeridas = [
        columna_coordinacion,
        'Código recuperador',
        'Nombre recuperador',
        'Cantidad Prestada',
        'Saldo capital',
        'Saldo vencido',
        'Saldo total',
        'Saldo riesgo capital',
        'Saldo riesgo total',
        '% MORA',
        'Días de mora'
    ]
    
    # Verificar y calcular columnas de riesgo si no existen
    if 'Saldo riesgo capital' not in df_completo.columns:
        logger.info("🔍 Calculando 'Saldo riesgo capital' (no existe en df_completo)")
        df_completo['Saldo riesgo capital'] = df_completo.apply(
            lambda row: row['Saldo capital'] if pd.notna(row.get('Días de mora', 0)) and row.get('Días de mora', 0) > 0 else 0,
            axis=1
        )
    
    if 'Saldo riesgo total' not in df_completo.columns:
        logger.info("🔍 Calculando 'Saldo riesgo total' (no existe en df_completo)")
        df_completo['Saldo riesgo total'] = df_completo.apply(
            lambda row: row['Saldo total'] if pd.notna(row.get('Días de mora', 0)) and row.get('Días de mora', 0) > 0 else 0,
            axis=1
        )
    
    if '% MORA' not in df_completo.columns:
        logger.info("🔍 Calculando '% MORA' (no existe en df_completo)")
        df_completo['% MORA'] = df_completo.apply(
            lambda row: (row['Saldo vencido'] / row['Saldo total']) 
                        if pd.notna(row.get('Saldo total', 0)) and row.get('Saldo total', 0) != 0 
                        else 0,
            axis=1
        )
    
    # Verificar columnas requeridas después de calcular las que faltaban
    # Las columnas de recuperador pueden no existir, así que las verificamos por separado
    columnas_faltantes = []
    for col in columnas_requeridas:
        if col not in df_completo.columns:
            # Si es una columna de recuperador, intentar variaciones
            if 'recuperador' in col.lower():
                # Buscar variaciones posibles
                posibles_nombres = [c for c in df_completo.columns if 'recuperador' in c.lower()]
                if not posibles_nombres:
                    columnas_faltantes.append(col)
                else:
                    logger.info(f"🔍 Columna '{col}' no encontrada, pero se encontraron variaciones: {posibles_nombres}")
            else:
                columnas_faltantes.append(col)
    
    if columnas_faltantes:
        logger.warning(f"⚠️ Columnas faltantes para X_Recuperador después de calcular: {columnas_faltantes}")
        logger.warning(f"🔍 Columnas disponibles en df_completo: {list(df_completo.columns)[:30]}")
        logger.warning(f"🔍 Total columnas: {len(df_completo.columns)}")
        # Si faltan columnas críticas (no recuperador), retornar vacío
        columnas_criticas = [col for col in columnas_faltantes if 'recuperador' not in col.lower()]
        if columnas_criticas:
            return pd.DataFrame()
        # Si solo faltan columnas de recuperador, continuar pero usar valores por defecto
    
    logger.info(f"✅ Columnas requeridas verificadas. Total registros: {len(df_completo)}")
    
    # Verificar si existen columnas de recuperador, si no, crear columnas dummy
    codigo_rec_col = 'Código recuperador' if 'Código recuperador' in df_completo.columns else None
    nombre_rec_col = 'Nombre recuperador' if 'Nombre recuperador' in df_completo.columns else None
    
    if codigo_rec_col is None or nombre_rec_col is None:
        logger.warning("⚠️ Columnas de recuperador no encontradas, usando valores por defecto")
        df_completo = df_completo.copy()
        if codigo_rec_col is None:
            df_completo['Código recuperador'] = 'N/A'
            codigo_rec_col = 'Código recuperador'
        if nombre_rec_col is None:
            df_completo['Nombre recuperador'] = 'N/A'
            nombre_rec_col = 'Nombre recuperador'
    
    # Agrupar por Coordinación + Recuperador y calcular agregaciones
    # Manejar valores NaN en las columnas de agrupación
    grupo = df_completo.groupby([columna_coordinacion, codigo_rec_col, nombre_rec_col], dropna=False).agg({
        'Cantidad Prestada': 'sum',
        'Saldo capital': 'sum',
        'Saldo vencido': 'sum',
        'Saldo total': 'sum',
        'Saldo riesgo capital': 'sum',
        'Saldo riesgo total': 'sum',
        '% MORA': 'mean'  # Promedio del porcentaje (se recalculará después)
    }).reset_index()
    
    # Calcular % MORA correcto (promedio ponderado o recálculo)
    # % MORA = Saldo vencido / Saldo total (por coordinación + recuperador)
    grupo['% MORA'] = grupo.apply(
        lambda row: (row['Saldo vencido'] / row['Saldo total']) 
                    if pd.notna(row['Saldo total']) and row['Saldo total'] != 0 
                    else 0,
        axis=1
    )
    
    # Calcular rangos de días de mora para cada coordinación + recuperador
    def calcular_rangos_mora(df_group, columna_mora='Días de mora', columna_riesgo='Saldo riesgo total'):
        """Calcula la suma de Saldo riesgo total por rangos de días de mora"""
        rangos = {
            '0': 0,
            '1-7': 0,
            '8-15': 0,
            '16-30': 0,
            '31-60': 0,
            '61-90': 0,
            'Mayor_90': 0
        }
        
        if columna_mora not in df_group.columns or columna_riesgo not in df_group.columns:
            return rangos
        
        for idx, row in df_group.iterrows():
            dias_mora = row[columna_mora] if pd.notna(row[columna_mora]) else 0
            saldo_riesgo = row[columna_riesgo] if pd.notna(row[columna_riesgo]) else 0
            
            if dias_mora == 0:
                rangos['0'] += saldo_riesgo
            elif 1 <= dias_mora <= 7:
                rangos['1-7'] += saldo_riesgo
            elif 8 <= dias_mora <= 15:
                rangos['8-15'] += saldo_riesgo
            elif 16 <= dias_mora <= 30:
                rangos['16-30'] += saldo_riesgo
            elif 31 <= dias_mora <= 60:
                rangos['31-60'] += saldo_riesgo
            elif 61 <= dias_mora <= 90:
                rangos['61-90'] += saldo_riesgo
            else:
                rangos['Mayor_90'] += saldo_riesgo
        
        return rangos
    
    # Calcular rangos para cada coordinación + recuperador
    rangos_por_grupo = []
    for idx, row in grupo.iterrows():
        coord = row[columna_coordinacion]
        codigo_rec = row[codigo_rec_col]
        nombre_rec = row[nombre_rec_col]
        
        # Filtrar datos del grupo específico (manejar NaN)
        mask = (df_completo[columna_coordinacion] == coord)
        if pd.notna(codigo_rec):
            mask = mask & (df_completo[codigo_rec_col] == codigo_rec)
        else:
            mask = mask & (df_completo[codigo_rec_col].isna())
        
        if pd.notna(nombre_rec):
            mask = mask & (df_completo[nombre_rec_col] == nombre_rec)
        else:
            mask = mask & (df_completo[nombre_rec_col].isna())
        
        df_group = df_completo[mask]
        
        rangos = calcular_rangos_mora(df_group)
        rangos_por_grupo.append(rangos)
    
    # Agregar columnas de rangos al grupo
    grupo['Rango_0'] = [r['0'] for r in rangos_por_grupo]
    grupo['Rango_1-7'] = [r['1-7'] for r in rangos_por_grupo]
    grupo['Rango_8-15'] = [r['8-15'] for r in rangos_por_grupo]
    grupo['Rango_16-30'] = [r['16-30'] for r in rangos_por_grupo]
    grupo['Rango_31-60'] = [r['31-60'] for r in rangos_por_grupo]
    grupo['Rango_61-90'] = [r['61-90'] for r in rangos_por_grupo]
    grupo['Rango_Mayor_90'] = [r['Mayor_90'] for r in rangos_por_grupo]
    
    # Calcular total general
    total_general = {
        columna_coordinacion: 'Total',
        codigo_rec_col: '',
        nombre_rec_col: '',
        'Cantidad Prestada': grupo['Cantidad Prestada'].sum(),
        'Saldo capital': grupo['Saldo capital'].sum(),
        'Saldo vencido': grupo['Saldo vencido'].sum(),
        'Saldo total': grupo['Saldo total'].sum(),
        'Saldo riesgo capital': grupo['Saldo riesgo capital'].sum(),
        'Saldo riesgo total': grupo['Saldo riesgo total'].sum(),
        '% MORA': (grupo['Saldo vencido'].sum() / grupo['Saldo total'].sum()) 
                  if grupo['Saldo total'].sum() != 0 else 0,
        'Rango_0': grupo['Rango_0'].sum(),
        'Rango_1-7': grupo['Rango_1-7'].sum(),
        'Rango_8-15': grupo['Rango_8-15'].sum(),
        'Rango_16-30': grupo['Rango_16-30'].sum(),
        'Rango_31-60': grupo['Rango_31-60'].sum(),
        'Rango_61-90': grupo['Rango_61-90'].sum(),
        'Rango_Mayor_90': grupo['Rango_Mayor_90'].sum()
    }
    
    # Crear DataFrame final con estructura específica
    df_resultado = grupo.copy()
    
    # Renombrar columnas para que coincidan con el formato esperado
    rename_dict = {
        columna_coordinacion: 'Coordinación',
        'Rango_0': 'Rango_0',
        'Rango_1-7': 'Rango_1-7',
        'Rango_8-15': 'Rango_8-15',
        'Rango_16-30': 'Rango_16-30',
        'Rango_31-60': 'Rango_31-60',
        'Rango_61-90': 'Rango_61-90',
        'Rango_Mayor_90': 'Rango_Mayor_90'
    }
    # Asegurar que las columnas de recuperador tengan los nombres correctos
    if codigo_rec_col != 'Código recuperador':
        rename_dict[codigo_rec_col] = 'Código recuperador'
    if nombre_rec_col != 'Nombre recuperador':
        rename_dict[nombre_rec_col] = 'Nombre recuperador'
    
    df_resultado = df_resultado.rename(columns=rename_dict)
    
    # Agregar fila de total general
    fila_total = pd.DataFrame([total_general])
    df_resultado = pd.concat([df_resultado, fila_total], ignore_index=True)
    
    logger.info(f"✅ Hoja X_Recuperador creada con {len(grupo)} grupos + 1 total")
    
    return df_resultado

def aplicar_formato_texto_concepto_deposito(worksheet, df):
    """
    Aplica formato de texto a la columna 'Concepto Depósito' para preservar ceros a la izquierda
    """
    if 'Concepto Depósito' in df.columns:
        for col_idx in range(1, worksheet.max_column + 1):
            if worksheet.cell(row=2, column=col_idx).value == 'Concepto Depósito':
                for row in range(3, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=col_idx).number_format = '@'
                logger.info(f"✅ Formato de texto aplicado a columna 'Concepto Depósito' (columna {col_idx})")
                break

def aplicar_formato_porcentaje_mora(worksheet, df):
    """
    Aplica formato de porcentaje a la columna '% MORA' (formato de porcentaje 0-100%)
    Excel automáticamente multiplicará los valores (que están entre 0-1) por 100 para mostrarlos como porcentaje
    """
    if '% MORA' in df.columns:
        for col_idx in range(1, worksheet.max_column + 1):
            if worksheet.cell(row=2, column=col_idx).value == '% MORA':
                for row in range(3, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=col_idx).number_format = '0.00%'  # Formato de porcentaje con 2 decimales
                logger.info(f"✅ Formato de porcentaje aplicado a columna '% MORA' (columna {col_idx})")
                break

def aplicar_formato_alerta(worksheet, df):
    """Aplica relleno rojo suave a celdas de columna 'Alerta' con valor 1."""
    if 'Alerta' not in df.columns:
        return
    alert_fill = PatternFill(start_color=COLORS.get('alert_red', 'F4CCCC'), end_color=COLORS.get('alert_red', 'F4CCCC'), fill_type='solid')
    for col_idx in range(1, worksheet.max_column + 1):
        if worksheet.cell(row=2, column=col_idx).value == 'Alerta':
            for row in range(3, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                if cell.value == 1 or cell.value == 1.0:
                    cell.fill = alert_fill
            logger.info(f"✅ Formato rojo suave aplicado a columna 'Alerta' (columna {col_idx})")
            break

def aplicar_formato_final(worksheet, df, es_hoja_mora=False):
    """Autoajuste de columnas, formato de moneda, fecha corta, y formatos especiales."""
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # a) Autoajuste de columnas
    for i in range(1, worksheet.max_column + 1):
        column_letter = get_column_letter(i)
        max_length = 0
        for j in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=j, column=i).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # b) Formato de encabezados (Fila 2)
    worksheet.row_dimensions[2].height = EXCEL_CONFIG['header_height']
    for cell in worksheet[2]:
        cell.font = Font(bold=True)
    
    # c) Relleno azul en "Días de mora" (en todas las hojas)
    for cell in worksheet[2]:
        if cell.value == 'Días de mora':
            cell.fill = PatternFill(start_color=COLORS['light_blue'], end_color=COLORS['light_blue'], fill_type="solid")
            break

    # d) Formato de moneda en columnas conocidas del df
    # Excluir columnas de días (que pueden contener "pago" en su nombre pero son numéricas enteras)
    COLUMNAS_NO_MONEDA = {'días desde el último pago', 'dias desde el ultimo pago'}
    columnas_moneda = [
        col for col in df.columns
        if any(key in col.lower() for key in CURRENCY_COLUMNS_KEYWORDS)
        and col.lower().strip() not in COLUMNAS_NO_MONEDA
    ]
    for col_name in columnas_moneda:
        if col_name in df.columns:
            # Buscar índice de la columna por encabezado (fila 2)
            for cell in worksheet[2]:
                if cell.value == col_name:
                    col_idx = cell.column
                    # Aplicar formato desde fila 3 (datos)
                    for row in range(3, worksheet.max_row + 1):
                        worksheet.cell(row=row, column=col_idx).number_format = EXCEL_CONFIG['currency_format']
                    break

    # e) Formato de fecha corta para columnas datetime del df
    columnas_fecha = df.select_dtypes(include=['datetime64[ns]', 'datetime64[ns, UTC]']).columns.tolist()
    for col_name in columnas_fecha:
        # Ubicar columna
        for cell in worksheet[2]:
            if cell.value == col_name:
                col_idx = cell.column
                for row in range(3, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=col_idx).number_format = EXCEL_CONFIG['date_format']
                break

    # f) Relleno azul en encabezados específicos de la hoja "Mora"
    if es_hoja_mora:
        for col_name in MORA_BLUE_COLUMNS:
            if col_name in df.columns:
                # Buscar y aplicar relleno azul al encabezado (fila 2)
                for cell in worksheet[2]:
                    if cell.value == col_name:
                        cell.fill = PatternFill(start_color=COLORS['light_blue'], end_color=COLORS['light_blue'], fill_type="solid")
                        break

    # g) Inmovilización de paneles en A3
    worksheet.freeze_panes = EXCEL_CONFIG['freeze_panes']

def aplicar_formato_condicional(worksheet, columna_mora, num_filas):
    """Aplica formato condicional de colores a la columna de días de mora"""
    color_scale_rule = ColorScaleRule(
        start_type='min', start_color='7AB800', # Verde
        mid_type='percentile', mid_value=50, mid_color='FFEB84', # Amarillo
        end_type='max', end_color='FF6464' # Rojo
    )
    
    # Encuentra la letra de la columna 'Días de mora' (ahora en fila 2)
    mora_col_letter = [col[0].column_letter for col in worksheet.iter_cols(min_row=2, max_row=2) if col[0].value == columna_mora][0]
    # Aplicar formato desde fila 3 (datos) hasta el final
    worksheet.conditional_formatting.add(f'{mora_col_letter}3:{mora_col_letter}{num_filas + 2}', color_scale_rule)

def crear_tabla_excel(worksheet, df, sheet_name, incluir_columnas_adicionales=False):
    """
    Convierte un rango de datos en una tabla formal de Excel.
    
    Si incluir_columnas_adicionales=True (solo para hoja Mora):
    - Fila 1: Títulos superiores combinados y coloreados
    - Fila 2: Encabezados de la tabla (datos + 9 columnas adicionales)
    - Fila 3+: Datos de la tabla
    
    Si incluir_columnas_adicionales=False (otras hojas):
    - Fila 2: Encabezados de la tabla (solo datos del DataFrame)
    - Fila 3+: Datos de la tabla
    
    Args:
        worksheet: El objeto de la hoja de Excel
        df: El DataFrame correspondiente con los datos
        sheet_name: El nombre de la hoja para crear un nombre único de tabla
        incluir_columnas_adicionales: Si True, incluye 9 columnas adicionales con títulos (solo para Mora)
    """
    try:
        from openpyxl.styles import PatternFill, Alignment
        
        # Calcular el rango de la tabla dinámicamente
        num_filas_datos = len(df)  # Número de filas de datos
        num_filas_tabla = num_filas_datos + 1  # +1 para incluir el encabezado
        num_columnas_originales = len(df.columns)
        
        # Obtener las letras de las columnas
        col_inicio = get_column_letter(1)  # Columna A
        col_fin_original = get_column_letter(num_columnas_originales)  # Última columna de datos
        
        if incluir_columnas_adicionales:
            # Lógica para hoja Mora: incluir 9 columnas adicionales
            num_columnas_totales = num_columnas_originales + 9  # +9 columnas adicionales
            col_fin_total = get_column_letter(num_columnas_totales)  # Última columna incluyendo las 9 adicionales
            
            # --- PASO 1: Crear títulos combinados con colores en la FILA 1 ---
            # Título 1 (Verde): "Seguimiento Call Center" - primeras 4 columnas adicionales
            titulo1_inicio = get_column_letter(num_columnas_originales + 1)
            titulo1_fin = get_column_letter(num_columnas_originales + ADDITIONAL_COLUMNS['titles']['green']['columns'])
            rango_titulo1 = f"{titulo1_inicio}1:{titulo1_fin}1"
            worksheet.merge_cells(rango_titulo1)
            
            celda_titulo1 = worksheet[f"{titulo1_inicio}1"]
            celda_titulo1.value = ADDITIONAL_COLUMNS['titles']['green']['text']
            celda_titulo1.fill = PatternFill(start_color=COLORS['green'], end_color=COLORS['green'], fill_type="solid")
            celda_titulo1.alignment = Alignment(horizontal="center", vertical="center")
            
            # Título 2 (Azul): "Gestión de Cobranza en Campo" - siguientes 5 columnas adicionales
            titulo2_inicio = get_column_letter(num_columnas_originales + ADDITIONAL_COLUMNS['titles']['green']['columns'] + 1)
            titulo2_fin = get_column_letter(num_columnas_originales + ADDITIONAL_COLUMNS['count'])
            rango_titulo2 = f"{titulo2_inicio}1:{titulo2_fin}1"
            worksheet.merge_cells(rango_titulo2)
            
            celda_titulo2 = worksheet[f"{titulo2_inicio}1"]
            celda_titulo2.value = ADDITIONAL_COLUMNS['titles']['blue']['text']
            celda_titulo2.fill = PatternFill(start_color=COLORS['blue'], end_color=COLORS['blue'], fill_type="solid")
            celda_titulo2.alignment = Alignment(horizontal="center", vertical="center")
            
            # --- PASO 2: Agregar encabezados específicos en la FILA 2 ---
            for i, encabezado in enumerate(ADDITIONAL_COLUMNS['headers']):
                col_letra = get_column_letter(num_columnas_originales + 1 + i)
                worksheet.cell(row=2, column=num_columnas_originales + 1 + i, value=encabezado)
            
            # Crear el rango de la tabla incluyendo las 9 columnas adicionales, empezando en fila 2
            rango_tabla = f"{col_inicio}2:{col_fin_total}{num_filas_tabla + 1}"  # +1 porque startrow=1 mueve todo una fila
        else:
            # Lógica para otras hojas: solo datos del DataFrame, terminar con "Criticidad"
            rango_tabla = f"{col_inicio}2:{col_fin_original}{num_filas_tabla + 1}"  # +1 porque startrow=1 mueve todo una fila
        
        # Crear el objeto Table con nombre único y válido
        nombre_tabla = generate_valid_table_name(sheet_name)
        logger.info(f"Creando tabla con nombre: '{nombre_tabla}' para hoja: '{sheet_name}'")
        tabla = Table(displayName=nombre_tabla, ref=rango_tabla)
        
        # Aplicar estilo a la tabla
        estilo = TableStyleInfo(
            name=EXCEL_CONFIG['table_style'],
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=False,
            showColumnStripes=False
        )
        tabla.tableStyleInfo = estilo
        
        # Añadir la tabla a la hoja
        worksheet.add_table(tabla)
        
        # Ajustar automáticamente el ancho de las columnas para que se vea todo el texto
        num_columnas_ajustar = num_columnas_totales if incluir_columnas_adicionales else num_columnas_originales
        fila_inicio_revision = 1 if incluir_columnas_adicionales else 2  # Revisar desde fila 1 si hay títulos, sino desde fila 2
        
        for i in range(1, num_columnas_ajustar + 1):
            column_letter = get_column_letter(i)
            max_length = 0
            # Revisar desde la fila apropiada hasta el final de los datos
            for row in range(fila_inicio_revision, num_filas_tabla + 2):  # +2 porque los datos empiezan en fila 2
                cell_value = worksheet.cell(row=row, column=i).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres de ancho
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
    except Exception as e:
        # Si hay algún error, no interrumpir el proceso principal
        logger.warning(f"No se pudo crear la tabla para la hoja {sheet_name}: {str(e)}")

def procesar_reporte_antiguedad(archivo_path, codigos_a_excluir=None):
    """Procesa el reporte de antigüedad con mejoras de robustez y mantenibilidad
    
    Args:
        archivo_path: Ruta del archivo Excel a procesar
        codigos_a_excluir: Lista opcional de códigos de acreditado a excluir del reporte
    """
    try:
        # Validar archivo
        validate_file_size(archivo_path)
        
        # --- PASO 1: Cargar y limpiar ---
        logger.info(f"Iniciando procesamiento del archivo: {archivo_path}")
        df = pd.read_excel(archivo_path, engine='openpyxl', dtype=DTYPE_CONFIG, header=0)
        df = clean_dataframe_columns(df)
        
        # Aplicar filtro de exclusión si se especifica
        if codigos_a_excluir:
            registros_antes = len(df)
            df = df[~df['Código acreditado'].isin(codigos_a_excluir)]
            logger.info(f"🔍 Filtro aplicado: Excluidos códigos {codigos_a_excluir}. Registros: {registros_antes} → {len(df)}")
        
        # Debug: Verificar las primeras filas después de la carga
        logger.info(f"🔍 DEBUG CARGA DE DATOS:")
        logger.info(f"   - Filas cargadas: {len(df)}")
        logger.info(f"   - Columnas: {list(df.columns)[:10]}...")  # Primeras 10 columnas
        if len(df) > 0:
            logger.info(f"   - Primera fila: {df.iloc[0].to_dict()}")
        
        
        # Obtener nombres de columnas desde configuración
        columna_codigo = COLUMN_MAPPING['codigo']
        columna_mora = COLUMN_MAPPING['mora']
        columna_coordinacion = COLUMN_MAPPING['coordinacion']
        columna_geolocalizacion = COLUMN_MAPPING['geolocalizacion']
        
        # Validar que las columnas requeridas existan
        required_columns = [columna_codigo, columna_mora, columna_coordinacion]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Columnas requeridas no encontradas: {missing_columns}")
        
        # --- PASO 1.1: Estandarizar códigos ANTES del filtrado ---
        # Estandarizar códigos a 6 dígitos para asegurar comparación correcta
        code_columns = [columna_codigo, 'Código promotor', 'Código recuperador']
        df = standardize_codes(df, code_columns)
        
        # --- PASO 2: Filtrar fraudes INMEDIATAMENTE después de la limpieza ---
        logger.info(f"Filtrando {len(LISTA_FRAUDE)} códigos de fraude")
        registros_antes_filtrado = len(df)
        df_filtrado = df[~df[columna_codigo].isin(LISTA_FRAUDE)]
        registros_eliminados = registros_antes_filtrado - len(df_filtrado)
        logger.info(f"Se eliminaron {registros_eliminados} registros por códigos fraudulentos")
        
        df_recup_000124_raw = None
        df_recup_000124_sin_links = None
        df_recup_000124_completo = None
        
        # --- PASO 2.1: Filtrar por códigos de recuperador a excluir (y guardar subset para hoja RECUPERADOR_000124) ---
        if CODIGOS_RECUPERADOR_EXCLUIR and 'Código recuperador' in df_filtrado.columns:
            registros_antes_recup = len(df_filtrado)
            codigos_excluir_norm = set()
            for c in CODIGOS_RECUPERADOR_EXCLUIR:
                s = str(c).strip()
                if s.replace('.', '').replace('-', '').isdigit():
                    codigos_excluir_norm.add(str(int(float(s))).zfill(6))
                else:
                    codigos_excluir_norm.add(s)
            mask_recup_excluir = df_filtrado['Código recuperador'].astype(str).str.zfill(6).isin(codigos_excluir_norm)
            df_recup_000124_raw = df_filtrado[mask_recup_excluir].copy()
            df_filtrado = df_filtrado[~mask_recup_excluir]
            eliminados_recup = registros_antes_recup - len(df_filtrado)
            if eliminados_recup > 0:
                logger.info(f"🔍 Filtro por recuperador: Excluidos códigos {CODIGOS_RECUPERADOR_EXCLUIR}. Registros: {registros_antes_recup} → {len(df_filtrado)} ({eliminados_recup} eliminados)")
        
        # Verificación de integridad de datos - ANTES de transformaciones (sobre datos filtrados)
        medio_comunic_1_antes = df_filtrado['Medio comunic. 1'].notna().sum() if 'Medio comunic. 1' in df_filtrado.columns else 0
        medio_comunic_2_antes = df_filtrado['Medio comunic. 2'].notna().sum() if 'Medio comunic. 2' in df_filtrado.columns else 0
        logger.info(f"Verificación de integridad - ANTES: 'Medio comunic. 1' -> {medio_comunic_1_antes}, 'Medio comunic. 2' -> {medio_comunic_2_antes}")
        
        # --- PASO 1.2: Limpieza de datos sobre DataFrame filtrado ---
        # Limpiar números de teléfono
        df_filtrado = clean_phone_numbers(df_filtrado)
        
        # Formatear columna Ciclo a 2 dígitos (01, 02, etc.)
        if 'Ciclo' in df_filtrado.columns:
            # Convertir a numérico primero, luego a string con 2 dígitos, rellenando con ceros a la izquierda
            df_filtrado['Ciclo'] = pd.to_numeric(df_filtrado['Ciclo'], errors='coerce').fillna(0).astype(int).astype(str).str.zfill(2)
            logger.info("✅ Columna 'Ciclo' formateada a 2 dígitos (01, 02, etc.)")
        
        # Añadir enlaces de geolocalización
        df_filtrado = add_geolocation_links(df_filtrado, columna_geolocalizacion)
        
        # DIAGNÓSTICO: Verificar columnas PAR en df_filtrado
        columnas_par_filtrado = [col for col in df_filtrado.columns if 'par' in str(col).lower()]
        if columnas_par_filtrado:
            logger.warning(f"🚨 PROBLEMA: df_filtrado tiene columnas PAR: {columnas_par_filtrado}")
        else:
            logger.info(f"✅ df_filtrado NO tiene columnas PAR")

        # --- PASO 1.3: Crear Informe Completo (después del filtrado) ---
        logger.info("Creando informe completo con registros filtrados")
        df_completo = df_filtrado.sort_values(by=columna_mora, ascending=False).copy()
        logger.info(f"🔍 df_completo ANTES de add_par_column: {[col for col in df_completo.columns if 'par' in str(col).lower()]}")
        df_completo = add_par_column(df_completo, columna_mora)
        
        # Insertar columna 'Link de Geolocalización' después de 'Geolocalización domicilio' si existen los links
        if 'link_texto' in df_completo.columns and columna_geolocalizacion in df_completo.columns:
            geo_index = df_completo.columns.get_loc(columna_geolocalizacion)
            df_completo.insert(geo_index + 1, 'Link de Geolocalización', df_completo['link_texto'])
            logger.info(f"📍 Insertada columna 'Link de Geolocalización' en df_completo después de '{columna_geolocalizacion}'")
        
        # Crear DataFrame sin las columnas temporales de links para escritura en Excel
        df_completo_sin_links = df_completo.drop(columns=['link_texto', 'link_url'], errors='ignore')

        # Reordenar columnas para que 'Código acreditado' sea la primera (solo en reporte completo)
        if 'Código acreditado' in df_completo_sin_links.columns:
            columnas = df_completo_sin_links.columns.tolist()
            # Remover 'Código acreditado' de su posición actual
            columnas.remove('Código acreditado')
            # Insertar 'Código acreditado' al inicio
            columnas.insert(0, 'Código acreditado')
            # Reordenar el DataFrame
            df_completo_sin_links = df_completo_sin_links[columnas]
            logger.info(f"📋 Reordenadas columnas en reporte completo: 'Código acreditado' es la primera columna")
        
        # DIAGNÓSTICO: Verificar si hay columnas duplicadas
        columnas_duplicadas = df_completo_sin_links.columns[df_completo_sin_links.columns.duplicated()].tolist()
        if columnas_duplicadas:
            logger.error(f"🚨 COLUMNAS DUPLICADAS encontradas: {columnas_duplicadas}")
            # Eliminar columnas duplicadas
            df_completo_sin_links = df_completo_sin_links.loc[:, ~df_completo_sin_links.columns.duplicated()]
            logger.info(f"🗑️ Columnas duplicadas eliminadas")

        # --- Pipeline para hoja RECUPERADOR_000124 (misma estructura que informe completo) ---
        if df_recup_000124_raw is not None and len(df_recup_000124_raw) > 0:
            dr = clean_phone_numbers(df_recup_000124_raw.copy())
            if 'Ciclo' in dr.columns:
                dr['Ciclo'] = pd.to_numeric(dr['Ciclo'], errors='coerce').fillna(0).astype(int).astype(str).str.zfill(2)
            dr = add_geolocation_links(dr, columna_geolocalizacion)
            dr = dr.sort_values(by=columna_mora, ascending=False).copy()
            dr = add_par_column(dr, columna_mora)
            if 'link_texto' in dr.columns and columna_geolocalizacion in dr.columns:
                geo_idx = dr.columns.get_loc(columna_geolocalizacion)
                dr.insert(geo_idx + 1, 'Link de Geolocalización', dr['link_texto'])
            df_recup_000124_completo = dr.copy()
            dr = dr.drop(columns=['link_texto', 'link_url'], errors='ignore')
            if 'Código acreditado' in dr.columns:
                cols = dr.columns.tolist()
                cols.remove('Código acreditado')
                cols.insert(0, 'Código acreditado')
                dr = dr[cols]
            dr = dr.loc[:, ~dr.columns.duplicated()] if dr.columns.duplicated().any() else dr
            dr = agregar_columna_concepto_deposito(dr.copy())
            dr = agregar_columnas_riesgo_y_mora(dr.copy())
            dr = agregar_columnas_dias_ultimo_pago_y_alerta(dr)
            df_recup_000124_sin_links = dr
            logger.info(f"📋 Preparados {len(df_recup_000124_sin_links)} registros para hoja RECUPERADOR_000124")

        # --- PASO 3: Ordenar y añadir columnas calculadas (sobre datos filtrados) ---
        df_ordenado = df_filtrado.sort_values(by=columna_mora, ascending=False).copy()
        logger.info(f"🔍 df_ordenado ANTES de add_par_column: {[col for col in df_ordenado.columns if 'par' in str(col).lower()]}")
        df_ordenado = add_par_column(df_ordenado, columna_mora)
        
        
        # Verificación de integridad de datos - DESPUÉS de transformaciones (sobre datos filtrados)
        medio_comunic_1_despues = df_ordenado['Medio comunic. 1'].notna().sum() if 'Medio comunic. 1' in df_ordenado.columns else 0
        medio_comunic_2_despues = df_ordenado['Medio comunic. 2'].notna().sum() if 'Medio comunic. 2' in df_ordenado.columns else 0
        
        # Verificar integridad
        if medio_comunic_1_antes == medio_comunic_1_despues:
            logger.info(f"Verificación 'Medio comunic. 1': Antes -> {medio_comunic_1_antes}, Después -> {medio_comunic_1_despues}. OK.")
        else:
            logger.warning(f"Verificación 'Medio comunic. 1': Antes -> {medio_comunic_1_antes}, Después -> {medio_comunic_1_despues}. PÉRDIDA DE DATOS!")
            
        if medio_comunic_2_antes == medio_comunic_2_despues:
            logger.info(f"Verificación 'Medio comunic. 2': Antes -> {medio_comunic_2_antes}, Después -> {medio_comunic_2_despues}. OK.")
        else:
            logger.warning(f"Verificación 'Medio comunic. 2': Antes -> {medio_comunic_2_antes}, Después -> {medio_comunic_2_despues}. PÉRDIDA DE DATOS!")

        # --- PASO 4: Crear DataFrame de Mora ---
        df_mora = df_ordenado[df_ordenado[columna_mora] >= 1].copy()
        logger.info(f"Registros en mora: {len(df_mora)}")
        
        # Aplicar add_par_column a df_mora para eliminar columnas duplicadas y regenerar 'PAR'
        logger.info(f"🔍 df_mora ANTES de add_par_column: {[col for col in df_mora.columns if 'par' in str(col).lower()]}")
        df_mora = add_par_column(df_mora, columna_mora)
        
        # Insertar columna 'Link de Geolocalización' después de 'Geolocalización domicilio' si existen los links
        if 'link_texto' in df_mora.columns and columna_geolocalizacion in df_mora.columns:
            geo_index = df_mora.columns.get_loc(columna_geolocalizacion)
            df_mora.insert(geo_index + 1, 'Link de Geolocalización', df_mora['link_texto'])
            logger.info(f"📍 Insertada columna 'Link de Geolocalización' en df_mora después de '{columna_geolocalizacion}'")
        
        # --- PASO 4.1: Crear DataFrame de Cuentas con Saldo Vencido ---
        columna_saldo_vencido = COLUMN_MAPPING.get('saldo_vencido', 'Saldo vencido')
        
        # Verificar si existe la columna 'Saldo vencido'
        if columna_saldo_vencido in df_ordenado.columns:
            # Crear filtro: Saldo vencido >= 1 Y (Días de mora <= 0 O nulo)
            df_saldo_vencido = df_ordenado[
                (df_ordenado[columna_saldo_vencido] >= 1) & 
                (pd.isna(df_ordenado[columna_mora]) | (df_ordenado[columna_mora] <= 0))
            ].copy()
            logger.info(f"Registros con saldo vencido >= 1 y sin mora: {len(df_saldo_vencido)}")
            
            if len(df_saldo_vencido) > 0:
                # Aplicar add_par_column a df_saldo_vencido para eliminar columnas duplicadas y regenerar 'PAR'
                logger.info(f"🔍 df_saldo_vencido ANTES de add_par_column: {[col for col in df_saldo_vencido.columns if 'par' in str(col).lower()]}")
                df_saldo_vencido = add_par_column(df_saldo_vencido, columna_mora)
                
                # Insertar columna 'Link de Geolocalización' después de 'Geolocalización domicilio' si existen los links
                if 'link_texto' in df_saldo_vencido.columns and columna_geolocalizacion in df_saldo_vencido.columns:
                    geo_index = df_saldo_vencido.columns.get_loc(columna_geolocalizacion)
                    df_saldo_vencido.insert(geo_index + 1, 'Link de Geolocalización', df_saldo_vencido['link_texto'])
                    logger.info(f"📍 Insertada columna 'Link de Geolocalización' en df_saldo_vencido después de '{columna_geolocalizacion}'")
            else:
                logger.info("No se encontraron registros con saldo vencido >= 1 y sin mora")
        else:
            logger.warning(f"⚠️ Columna '{columna_saldo_vencido}' no encontrada en DataFrame. Saltando creación de hoja 'Cuentas con saldo vencido'")
            df_saldo_vencido = None

        # --- PASO 5: Distribuir ---
        coordinaciones_data = {}
        lista_coordinaciones = df_ordenado[columna_coordinacion].unique()
        for coord in lista_coordinaciones:
            if pd.notna(coord):
                df_coord = df_ordenado[df_ordenado[columna_coordinacion] == coord].copy()
                
                # Aplicar add_par_column a df_coord para eliminar columnas duplicadas y regenerar 'PAR'
                logger.info(f"🔍 df_coord '{coord}' ANTES de add_par_column: {[col for col in df_coord.columns if 'par' in str(col).lower()]}")
                df_coord = add_par_column(df_coord, columna_mora)
                
                # Insertar columna 'Link de Geolocalización' después de 'Geolocalización domicilio' si existen los links
                if 'link_texto' in df_coord.columns and columna_geolocalizacion in df_coord.columns:
                    geo_index = df_coord.columns.get_loc(columna_geolocalizacion)
                    df_coord.insert(geo_index + 1, 'Link de Geolocalización', df_coord['link_texto'])
                    logger.info(f"📍 Insertada columna 'Link de Geolocalización' en coordinación '{coord}' después de '{columna_geolocalizacion}'")
                
                coordinaciones_data[coord] = df_coord
                

        # --- PASO 6: Generar el archivo Excel final ---
        # Calcular fecha del reporte: día anterior, excepto lunes que usa viernes
        hoy = datetime.now()
        if hoy.weekday() == 0:  # Lunes (0 = lunes)
            fecha_reporte = hoy - timedelta(days=3)  # Viernes anterior
        else:
            fecha_reporte = hoy - timedelta(days=1)  # Día anterior
        
        fecha_actual = fecha_reporte.strftime("%d%m%Y")
        nombre_archivo_salida = f'ReportedeAntigüedad_{fecha_actual}.xlsx'
        ruta_salida = os.path.join('uploads', nombre_archivo_salida)
        
        # Buscar plantilla con tablas dinámicas
        directorio_raiz = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        plantilla_path = os.path.join(directorio_raiz, "PLANTIILA2.xlsx")
        usar_plantilla = os.path.exists(plantilla_path)
        
        if usar_plantilla:
            # --- Flujo con plantilla: usar tablas dinámicas existentes ---
            logger.info(f"📋 Plantilla con tablas dinámicas encontrada: {plantilla_path}")
            import shutil
            shutil.copy(plantilla_path, ruta_salida)
            logger.info(f"📋 Plantilla copiada a: {ruta_salida}")
            
            # Abrir el archivo copiado con openpyxl para llenar R_Completo
            import openpyxl
            wb_plantilla = openpyxl.load_workbook(ruta_salida)
            
            # Preparar datos para R_Completo
            df_r_completo = df_completo_sin_links.copy()
            df_r_completo = agregar_columna_concepto_deposito(df_r_completo)
            df_r_completo = agregar_columnas_riesgo_y_mora(df_r_completo)
            df_r_completo = agregar_columnas_dias_ultimo_pago_y_alerta(df_r_completo)
            df_r_completo = agregar_columnas_nuevas(df_r_completo)

            # Llenar hoja R_Completo con los datos
            if 'R_Completo' in wb_plantilla.sheetnames:
                ws_r_completo = wb_plantilla['R_Completo']
                
                # Escribir encabezados en fila 2 (incluye "Días desde el último pago" y "Alerta")
                for col_idx, col_name in enumerate(df_r_completo.columns, start=1):
                    cell = ws_r_completo.cell(row=2, column=col_idx, value=col_name)
                    cell.font = Font(bold=True)
                ws_r_completo.row_dimensions[2].height = EXCEL_CONFIG['header_height']
                # Relleno azul en encabezado "Días de mora"
                for col_idx, col_name in enumerate(df_r_completo.columns, start=1):
                    if col_name == COLUMN_MAPPING.get('mora', 'Días de mora'):
                        ws_r_completo.cell(row=2, column=col_idx).fill = PatternFill(
                            start_color=COLORS['light_blue'], end_color=COLORS['light_blue'], fill_type='solid')
                        break
                
                # Escribir datos desde fila 3
                logger.info(f"📝 Escribiendo {len(df_r_completo)} filas en R_Completo...")
                for row_idx, (_, row) in enumerate(df_r_completo.iterrows(), start=3):
                    for col_idx, value in enumerate(row, start=1):
                        cell = ws_r_completo.cell(row=row_idx, column=col_idx)
                        if pd.isna(value):
                            cell.value = None
                        else:
                            cell.value = value
                
                # Hipervínculos en columna 'Link de Geolocalización'
                if 'Link de Geolocalización' in df_r_completo.columns and 'link_texto' in df_completo.columns:
                    link_col_r = df_r_completo.columns.get_loc('Link de Geolocalización') + 1
                    for i, (_, fila) in enumerate(df_completo.iterrows()):
                        row_num = i + 3
                        escribir_hipervinculo_excel(ws_r_completo, row_num, link_col_r, fila.get('link_texto'), fila.get('link_url'))

                # Formato condicional degradado en columna 'Días de mora'
                col_mora_nombre = COLUMN_MAPPING.get('mora', 'Días de mora')
                aplicar_formato_condicional(ws_r_completo, col_mora_nombre, len(df_r_completo))

                # Aplicar formatos de porcentaje (% MORA) y Alerta (relleno rojo)
                aplicar_formato_porcentaje_mora(ws_r_completo, df_r_completo)
                aplicar_formato_alerta(ws_r_completo, df_r_completo)
                
                # Actualizar rango de la tabla existente para abarcar todos los datos escritos
                num_filas_escritas = len(df_r_completo)
                num_cols = len(df_r_completo.columns)
                ultima_col_letra = get_column_letter(num_cols)
                ultima_fila = num_filas_escritas + 2  # fila 2 encabezado + filas de datos
                nuevo_rango = f"A2:{ultima_col_letra}{ultima_fila}"
                if hasattr(ws_r_completo, 'tables') and ws_r_completo.tables:
                    for t_name in list(ws_r_completo.tables.keys()):
                        ws_r_completo.tables[t_name].ref = nuevo_rango
                        logger.info(f"✅ Rango de tabla '{t_name}' en R_Completo actualizado a {nuevo_rango}")
                
                logger.info(f"✅ R_Completo llenado con {len(df_r_completo)} registros")
            else:
                logger.warning("⚠️ Hoja 'R_Completo' no encontrada en la plantilla")
            
            # Guardar cambios
            wb_plantilla.save(ruta_salida)
            wb_plantilla.close()
            
            # Configurar tablas dinámicas para que se actualicen automáticamente al abrir
            # NOTA: Por ahora desactivado para pruebas - el usuario puede activar refreshOnLoad manualmente en la plantilla
            logger.info("ℹ️ Para actualización automática de tablas dinámicas, configura 'Actualizar al abrir' en la plantilla")
            
            # Usar ExcelWriter en modo append para agregar las demás hojas
            writer = pd.ExcelWriter(ruta_salida, engine='openpyxl', mode='a', if_sheet_exists='new')
        else:
            # --- Flujo sin plantilla: crear todo desde cero ---
            logger.info(f"ℹ️ Plantilla no encontrada en: {plantilla_path}")
            logger.info("   Generando archivo sin tablas dinámicas")
            writer = pd.ExcelWriter(ruta_salida, engine='openpyxl')
        
        with writer:
            # --- PASO 6.0: Crear hojas X_Coordinación y X_Recuperador (solo sin plantilla) ---
            # Si usamos plantilla, estas hojas ya existen con tablas dinámicas
            crear_hojas_resumen = not usar_plantilla
            
            if crear_hojas_resumen:
                logger.info("Creando hoja 'X_Coordinación' (PRIMERA HOJA)...")
                try:
                    df_x_coordinacion = crear_hoja_x_coordinacion(df_completo)
                    logger.info(f"🔍 DataFrame X_Coordinación creado: {len(df_x_coordinacion)} filas, {len(df_x_coordinacion.columns)} columnas")
                    logger.info(f"🔍 Columnas en df_x_coordinacion: {list(df_x_coordinacion.columns)}")
                except Exception as e:
                    logger.error(f"❌ Error creando hoja X_Coordinación: {str(e)}")
                    import traceback
                    logger.error(traceback.format_exc())
                    df_x_coordinacion = pd.DataFrame()
            else:
                df_x_coordinacion = pd.DataFrame()  # Vacío para saltar el bloque siguiente
                logger.info("📋 Usando plantilla - X_Coordinación ya existe con tabla dinámica")
            
            if not df_x_coordinacion.empty:
                # Reorganizar columnas del DataFrame para que coincidan con la estructura
                columnas_principales = [
                    'Coordinación', 'Cantidad Prestada', 'Saldo capital', 'Saldo vencido',
                    'Saldo total', 'Saldo riesgo capital', 'Saldo riesgo total', '% MORA'
                ]
                
                # Verificar que las columnas existan y reordenar
                columnas_disponibles = [col for col in columnas_principales if col in df_x_coordinacion.columns]
                columnas_adicionales = [col for col in df_x_coordinacion.columns if col not in columnas_principales]
                
                # Reordenar DataFrame
                df_x_coordinacion_ordenado = df_x_coordinacion[columnas_disponibles + columnas_adicionales].copy()
                
                # Renombrar columnas de rangos para que coincidan
                mapeo_rangos = {
                    'Rango_0': '0',
                    'Rango_1-7': '1-7 días',
                    'Rango_8-15': '8-15 días',
                    'Rango_16-30': '16-30 días',
                    'Rango_31-60': '31-60 días',
                    'Rango_61-90': '61-90 días',
                    'Rango_Mayor_90': 'Mayor_90'
                }
                
                # Renombrar columnas de rangos
                for col_antigua, col_nueva in mapeo_rangos.items():
                    if col_antigua in df_x_coordinacion_ordenado.columns:
                        df_x_coordinacion_ordenado = df_x_coordinacion_ordenado.rename(columns={col_antigua: col_nueva})
                
                logger.info(f"🔍 Escribiendo hoja X_Coordinación con {len(df_x_coordinacion_ordenado)} filas")
                
                # Escribir DataFrame empezando en fila 9 (después de encabezados en fila 6 y filas 7-8 comprimidas)
                df_x_coordinacion_ordenado.to_excel(writer, sheet_name='X_Coordinación', index=False, startrow=9)
                ws_x_coord = writer.sheets['X_Coordinación']
                
                logger.info(f"✅ Hoja X_Coordinación creada en Excel. Filas: {ws_x_coord.max_row}, Columnas: {ws_x_coord.max_column}")
                
                # Escribir estructura completa (filas 1-7) según formato objetivo
                # Filas 1-4: Vacías (no hacer nada)
                
                # Color azul para encabezados (sombreado claro con letra azul fuerte)
                color_fondo_azul_claro = 'D9E1F2'  # Azul claro para fondo
                color_texto_azul_fuerte = '002060'  # Azul fuerte para texto
                
                # Fila 5: "PAR" centrado desde J5 hasta O5 con fondo azul
                ws_x_coord.merge_cells('J5:O5')
                cell_par = ws_x_coord.cell(row=5, column=10)  # Columna J (10)
                cell_par.value = 'PAR'
                cell_par.font = Font(bold=True, size=11, color=color_texto_azul_fuerte)
                cell_par.fill = PatternFill(start_color=color_fondo_azul_claro, end_color=color_fondo_azul_claro, fill_type='solid')
                cell_par.alignment = Alignment(horizontal='center', vertical='center')
                cell_par.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                
                # Fila 6: Encabezados de AMBAS tablas
                # TABLA 1: Columnas A-H (1-8) - Encabezados principales
                encabezados_tabla1 = [
                    'Coordinación', 'Cantidad\nPrestada', 'Saldo\nCapital', 'Saldo\nVencido',
                    'Saldo\nTotal', 'Saldo\nRiesgo Capital', 'Saldo\nRiesgo Total', '% MORA'
                ]
                for col_idx, encabezado in enumerate(encabezados_tabla1, start=1):
                    cell = ws_x_coord.cell(row=6, column=col_idx)
                    cell.value = encabezado
                    cell.font = Font(bold=True, size=11, color=color_texto_azul_fuerte)
                    cell.fill = PatternFill(start_color=color_fondo_azul_claro, end_color=color_fondo_azul_claro, fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
                
                # Columna I (9): Vacía - separador entre tablas (no hacer nada)
                
                # TABLA 2: Columnas J-S (10-19) - Copiar encabezados de la fila 10 (donde pandas los escribió) a la fila 6
                # Primero, leer los encabezados que pandas escribió en la fila 10
                # Limpiar todos los encabezados de la fila 10 (tanto de tabla 1 como tabla 2) y copiar solo los de tabla 2 a fila 6
                for col_idx in range(1, ws_x_coord.max_column + 1):
                    cell_fila10 = ws_x_coord.cell(row=10, column=col_idx)
                    if cell_fila10.value is not None:
                        # Si es del segundo segmento (columna 10 en adelante), copiar a la fila 6
                        if col_idx >= 10:
                            cell_fila6 = ws_x_coord.cell(row=6, column=col_idx)
                            cell_fila6.value = cell_fila10.value
                            cell_fila6.font = Font(bold=True, size=11, color=color_texto_azul_fuerte)
                            cell_fila6.fill = PatternFill(start_color=color_fondo_azul_claro, end_color=color_fondo_azul_claro, fill_type='solid')
                            cell_fila6.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell_fila6.border = Border(
                                left=Side(style='thin', color='000000'),
                                right=Side(style='thin', color='000000'),
                                top=Side(style='thin', color='000000'),
                                bottom=Side(style='thin', color='000000')
                            )
                        # Limpiar la celda de la fila 10 (solo si no está fusionada)
                        limpiar_celda_segura(cell_fila10)
                
                # Fila 7: Comprimir (altura mínima) - Solo algunos valores específicos
                # Columna 10: "Suma de Saldo riesgo total"
                cell_suma = ws_x_coord.cell(row=7, column=10)
                cell_suma.value = 'Suma de Saldo riesgo total'
                cell_suma.font = Font(bold=True, size=9, color=color_texto_azul_fuerte)
                cell_suma.fill = PatternFill(start_color=color_fondo_azul_claro, end_color=color_fondo_azul_claro, fill_type='solid')
                cell_suma.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Columna 11: "PAR"
                cell_par_fila7 = ws_x_coord.cell(row=7, column=11)
                cell_par_fila7.value = 'PAR'
                cell_par_fila7.font = Font(bold=True, size=9, color=color_texto_azul_fuerte)
                cell_par_fila7.fill = PatternFill(start_color=color_fondo_azul_claro, end_color=color_fondo_azul_claro, fill_type='solid')
                cell_par_fila7.alignment = Alignment(horizontal='center', vertical='center')
                
                # Fila 8: Comprimir (altura mínima) - Solo algunos valores específicos
                # Columna 1: "Coordinación" (repetir)
                cell_coord_fila8 = ws_x_coord.cell(row=8, column=1)
                cell_coord_fila8.value = 'Coordinación'
                cell_coord_fila8.font = Font(bold=True, size=9, color=color_texto_azul_fuerte)
                cell_coord_fila8.fill = PatternFill(start_color=color_fondo_azul_claro, end_color=color_fondo_azul_claro, fill_type='solid')
                cell_coord_fila8.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Columna 10: "Etiquetas de fila" (repetir)
                cell_etiquetas = ws_x_coord.cell(row=8, column=10)
                cell_etiquetas.value = 'Etiquetas de fila'
                cell_etiquetas.font = Font(bold=True, size=9, color=color_texto_azul_fuerte)
                cell_etiquetas.fill = PatternFill(start_color=color_fondo_azul_claro, end_color=color_fondo_azul_claro, fill_type='solid')
                cell_etiquetas.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Columna 11: "PAR" (repetir)
                cell_par_fila8 = ws_x_coord.cell(row=8, column=11)
                cell_par_fila8.value = 'PAR'
                cell_par_fila8.font = Font(bold=True, size=9, color=color_texto_azul_fuerte)
                cell_par_fila8.fill = PatternFill(start_color=color_fondo_azul_claro, end_color=color_fondo_azul_claro, fill_type='solid')
                cell_par_fila8.alignment = Alignment(horizontal='center', vertical='center')
                
                # Aplicar formato a datos (fila 11+)
                # Formato de moneda a columnas numéricas
                columnas_moneda = ['Cantidad Prestada', 'Saldo capital', 'Saldo vencido', 'Saldo total',
                                  'Saldo riesgo capital', 'Saldo riesgo total'] + list(mapeo_rangos.values())
                for col_name in columnas_moneda:
                    if col_name in df_x_coordinacion_ordenado.columns:
                        col_idx = df_x_coordinacion_ordenado.columns.get_loc(col_name) + 1
                        for row in range(11, ws_x_coord.max_row + 1):
                            cell = ws_x_coord.cell(row=row, column=col_idx)
                            if cell.value is not None:
                                cell.number_format = EXCEL_CONFIG['currency_format']
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                                cell.border = Border(
                                    left=Side(style='thin', color='000000'),
                                    right=Side(style='thin', color='000000'),
                                    top=Side(style='thin', color='000000'),
                                    bottom=Side(style='thin', color='000000')
                                )
                
                # Formato de porcentaje a % MORA - Solo encabezado en amarillo
                color_amarillo = 'FFFF00'  # Amarillo
                if '% MORA' in df_x_coordinacion_ordenado.columns:
                    col_idx = df_x_coordinacion_ordenado.columns.get_loc('% MORA') + 1
                    # Aplicar formato a encabezado en fila 6 (solo el encabezado en amarillo)
                    cell_header = ws_x_coord.cell(row=6, column=col_idx)
                    cell_header.fill = PatternFill(start_color=color_amarillo, end_color=color_amarillo, fill_type='solid')
                    # Aplicar formato a datos (fila 11+) - SIN fondo amarillo, solo formato de porcentaje
                    for row in range(11, ws_x_coord.max_row + 1):
                        cell = ws_x_coord.cell(row=row, column=col_idx)
                        if cell.value is not None:
                            cell.number_format = '0.00%'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            cell.border = Border(
                                left=Side(style='thin', color='000000'),
                                right=Side(style='thin', color='000000'),
                                top=Side(style='thin', color='000000'),
                                bottom=Side(style='thin', color='000000')
                            )
                
                # Formato a columna Coordinación
                if 'Coordinación' in df_x_coordinacion_ordenado.columns:
                    col_idx = df_x_coordinacion_ordenado.columns.get_loc('Coordinación') + 1
                    for row in range(10, ws_x_coord.max_row + 1):
                        cell = ws_x_coord.cell(row=row, column=col_idx)
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        cell.border = Border(
                            left=Side(style='thin', color='000000'),
                            right=Side(style='thin', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thin', color='000000')
                        )
                        # Resaltar fila de Total con azul claro (mismo que encabezados)
                        if cell.value == 'Total':
                            for col in range(1, ws_x_coord.max_column + 1):
                                total_cell = ws_x_coord.cell(row=row, column=col)
                                total_cell.fill = PatternFill(start_color=color_fondo_azul_claro, end_color=color_fondo_azul_claro, fill_type='solid')
                                total_cell.font = Font(bold=True, color=color_texto_azul_fuerte)
                                total_cell.border = Border(
                                    left=Side(style='thin', color='000000'),
                                    right=Side(style='thin', color='000000'),
                                    top=Side(style='thin', color='000000'),
                                    bottom=Side(style='thin', color='000000')
                                )
                
                # Ajustar ancho de columnas
                for col_idx in range(1, ws_x_coord.max_column + 1):
                    column_letter = get_column_letter(col_idx)
                    if col_idx != 9:  # Todas las columnas excepto I
                        max_length = 0
                        for row in range(1, ws_x_coord.max_row + 1):
                            cell = ws_x_coord.cell(row=row, column=col_idx)
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        ws_x_coord.column_dimensions[column_letter].width = min(max_length + 2, 20)
                
                # Columna I (9): Completamente oculta (ancho 0 y hidden) - Hacerlo al final
                column_letter_i = get_column_letter(9)
                ws_x_coord.column_dimensions[column_letter_i].width = 0.0  # 0.00 de ancho
                ws_x_coord.column_dimensions[column_letter_i].hidden = True  # Ocultar completamente
                # Limpiar todas las celdas de la columna I
                for row in range(1, ws_x_coord.max_row + 1):
                    cell = ws_x_coord.cell(row=row, column=9)
                    limpiar_celda_segura(cell)
                
                # Ajustar altura de filas - Fila 6 visible y larga, filas 7-10 con altura 0 (completamente invisibles)
                ws_x_coord.row_dimensions[6].height = 30  # Fila 6: Visible y larga (encabezados)
                ws_x_coord.row_dimensions[7].height = 0.0   # Fila 7: Altura 0.0 (completamente invisible)
                ws_x_coord.row_dimensions[8].height = 0.0   # Fila 8: Altura 0.0 (completamente invisible)
                ws_x_coord.row_dimensions[9].height = 0.0   # Fila 9: Altura 0.0 (completamente invisible)
                ws_x_coord.row_dimensions[10].height = 0.0   # Fila 10: Altura 0.0 (completamente invisible)
                
                # Ocultar filas 7, 8, 9 y 10 completamente (método adicional)
                ws_x_coord.row_dimensions[7].hidden = True
                ws_x_coord.row_dimensions[8].hidden = True
                ws_x_coord.row_dimensions[9].hidden = True
                ws_x_coord.row_dimensions[10].hidden = True
                
                # Determinar límites reales de la tabla
                # Última columna con datos: buscar en todas las filas de datos (11+)
                ultima_col_con_datos = 0
                ultima_fila_con_datos = 0
                for row in range(11, ws_x_coord.max_row + 1):
                    for col in range(1, ws_x_coord.max_column + 1):
                        cell = ws_x_coord.cell(row=row, column=col)
                        if cell.value is not None and str(cell.value).strip() != '':
                            ultima_col_con_datos = max(ultima_col_con_datos, col)
                            ultima_fila_con_datos = max(ultima_fila_con_datos, row)
                
                # Si no encontramos datos, usar la última columna del segundo segmento (S = 19)
                if ultima_col_con_datos == 0:
                    ultima_col_con_datos = 19  # Columna S
                
                logger.info(f"📊 Límites de tabla: Última columna={ultima_col_con_datos}, Última fila={ultima_fila_con_datos}")
                
                # Limpiar filas 1-4 (fuera del área) - completamente blancas sin bordes
                for row in range(1, 5):
                    for col in range(1, ws_x_coord.max_column + 1):
                        cell = ws_x_coord.cell(row=row, column=col)
                        limpiar_celda_segura(cell)
                
                # Limpiar fila 5 excepto la celda PAR (J5:O5)
                for col in range(1, ws_x_coord.max_column + 1):
                    # Si no es parte de PAR (J5:O5), limpiar
                    if not (col >= 10 and col <= 15):  # J=10, O=15
                        cell = ws_x_coord.cell(row=5, column=col)
                        limpiar_celda_segura(cell)
                
                # Limpiar columnas a la DERECHA de la tabla (después de la última columna con datos)
                for row in range(1, ws_x_coord.max_row + 1):
                    for col in range(ultima_col_con_datos + 1, ws_x_coord.max_column + 1):
                        cell = ws_x_coord.cell(row=row, column=col)
                        limpiar_celda_segura(cell)
                
                # Limpiar filas ABAJO de la tabla (después de la última fila con datos)
                if ultima_fila_con_datos > 0:
                    for row in range(ultima_fila_con_datos + 1, ws_x_coord.max_row + 1):
                        for col in range(1, ws_x_coord.max_column + 1):
                            cell = ws_x_coord.cell(row=row, column=col)
                            limpiar_celda_segura(cell)
                
                # Limpiar celdas vacías en las filas de datos (después de la última columna con datos)
                for row in range(11, ultima_fila_con_datos + 1 if ultima_fila_con_datos > 0 else ws_x_coord.max_row + 1):
                    for col in range(ultima_col_con_datos + 1, ws_x_coord.max_column + 1):
                        cell = ws_x_coord.cell(row=row, column=col)
                        limpiar_celda_segura(cell)
                
                # Limpiar también las celdas fuera del rango de datos en la fila 6 (encabezados)
                # Área principal: A-H (1-8) y J-S (10-19)
                for col in range(1, ws_x_coord.max_column + 1):
                    if col != 9:  # No tocar columna I (ya está oculta)
                        cell = ws_x_coord.cell(row=6, column=col)
                        # Si no está en el rango A-H o J-S, limpiar
                        if not ((col >= 1 and col <= 8) or (col >= 10 and col <= 19)):
                            limpiar_celda_segura(cell)
                
                # LIMPIEZA FINAL: Asegurar que TODAS las celdas fuera del área estén blancas sin bordes
                # Esto se hace al final para evitar que otros formatos sobrescriban
                # Área principal: filas 5-6 (encabezados) y filas 11 hasta ultima_fila_con_datos (datos)
                # Columnas: A-H (1-8) y J hasta ultima_col_con_datos (10+)
                
                # Limpiar TODAS las celdas fuera del área principal (hacerlo al final)
                for row in range(1, ws_x_coord.max_row + 1):
                    for col in range(1, ws_x_coord.max_column + 1):
                        # Determinar si la celda está dentro del área principal
                        en_area_principal = False
                        
                        # Fila 5: Solo PAR (J5:O5) está en el área
                        if row == 5:
                            if col >= 10 and col <= 15:  # J-O
                                en_area_principal = True
                        
                        # Fila 6: Columnas A-H y J-S están en el área
                        elif row == 6:
                            if (col >= 1 and col <= 8) or (col >= 10 and col <= 19):  # A-H o J-S
                                en_area_principal = True
                        
                        # Filas 11+: Columnas A-H y J hasta ultima_col_con_datos están en el área
                        elif row >= 11 and row <= ultima_fila_con_datos:
                            if (col >= 1 and col <= 8) or (col >= 10 and col <= ultima_col_con_datos):
                                en_area_principal = True
                        
                        # Si NO está en el área principal, limpiar completamente
                        if not en_area_principal:
                            cell = ws_x_coord.cell(row=row, column=col)
                            if limpiar_celda_segura(cell):
                                try:
                                    cell.font = Font()  # Resetear fuente
                                    cell.alignment = Alignment()  # Resetear alineación
                                except:
                                    pass
                
                logger.info("✅ Hoja 'X_Coordinación' creada exitosamente como PRIMERA HOJA")
            else:
                logger.warning("⚠️ No se pudo crear la hoja 'X_Coordinación' (DataFrame vacío)")
                logger.warning(f"🔍 Columnas disponibles en df_completo: {list(df_completo.columns)[:20]}...")
                logger.warning(f"🔍 Total columnas en df_completo: {len(df_completo.columns)}")
            
            # --- PASO 6.1: Crear hoja "X_Recuperador" SEGUNDA (solo sin plantilla) ---
            if crear_hojas_resumen:
                logger.info("Creando hoja 'X_Recuperador' (SEGUNDA HOJA)...")
                try:
                    df_x_recuperador = crear_hoja_x_recuperador(df_completo)
                    logger.info(f"🔍 DataFrame X_Recuperador creado: {len(df_x_recuperador)} filas, {len(df_x_recuperador.columns)} columnas")
                    logger.info(f"🔍 Columnas en df_x_recuperador: {list(df_x_recuperador.columns)}")
                except Exception as e:
                    logger.error(f"❌ Error creando hoja X_Recuperador: {str(e)}")
                    import traceback
                    logger.error(traceback.format_exc())
                    df_x_recuperador = pd.DataFrame()
            else:
                df_x_recuperador = pd.DataFrame()  # Vacío para saltar el bloque siguiente
                logger.info("📋 Usando plantilla - X_Recuperador ya existe con tabla dinámica")
            
            if not df_x_recuperador.empty:
                # Reorganizar columnas del DataFrame para que coincidan con la estructura
                # X_Recuperador tiene: Coordinación, Código recuperador, Nombre recuperador, luego las métricas
                columnas_principales = [
                    'Coordinación', 'Código recuperador', 'Nombre recuperador',
                    'Cantidad Prestada', 'Saldo capital', 'Saldo vencido',
                    'Saldo total', 'Saldo riesgo capital', 'Saldo riesgo total', '% MORA'
                ]
                
                # Verificar que las columnas existan y reordenar
                columnas_disponibles = [col for col in columnas_principales if col in df_x_recuperador.columns]
                columnas_adicionales = [col for col in df_x_recuperador.columns if col not in columnas_principales]
                
                # Reordenar DataFrame
                df_x_recuperador_ordenado = df_x_recuperador[columnas_disponibles + columnas_adicionales].copy()
                
                # Renombrar columnas de rangos para que coincidan
                mapeo_rangos = {
                    'Rango_0': '0',
                    'Rango_1-7': '1-7 días',
                    'Rango_8-15': '8-15 días',
                    'Rango_16-30': '16-30 días',
                    'Rango_31-60': '31-60 días',
                    'Rango_61-90': '61-90 días',
                    'Rango_Mayor_90': 'Mayor_90'
                }
                
                # Renombrar columnas de rangos
                for col_antigua, col_nueva in mapeo_rangos.items():
                    if col_antigua in df_x_recuperador_ordenado.columns:
                        df_x_recuperador_ordenado = df_x_recuperador_ordenado.rename(columns={col_antigua: col_nueva})
                
                logger.info(f"🔍 Escribiendo hoja X_Recuperador con {len(df_x_recuperador_ordenado)} filas")
                
                # Escribir DataFrame empezando en fila 9 (después de encabezados en fila 6 y filas 7-8 comprimidas)
                df_x_recuperador_ordenado.to_excel(writer, sheet_name='X_Recuperador', index=False, startrow=9)
                ws_x_recup = writer.sheets['X_Recuperador']
                
                logger.info(f"✅ Hoja X_Recuperador creada en Excel. Filas: {ws_x_recup.max_row}, Columnas: {ws_x_recup.max_column}")
                
                # Aplicar el mismo formato que X_Coordinación (reutilizar código)
                # Color azul para encabezados
                color_fondo_azul_claro = 'D9E1F2'
                color_texto_azul_fuerte = '002060'
                
                # Fila 5: "PAR" centrado desde J5 hasta O5 con fondo azul
                ws_x_recup.merge_cells('J5:O5')
                cell_par = ws_x_recup.cell(row=5, column=10)
                cell_par.value = 'PAR'
                cell_par.font = Font(bold=True, size=11, color=color_texto_azul_fuerte)
                cell_par.fill = PatternFill(start_color=color_fondo_azul_claro, end_color=color_fondo_azul_claro, fill_type='solid')
                cell_par.alignment = Alignment(horizontal='center', vertical='center')
                cell_par.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                
                # Fila 6: Encabezados de AMBAS tablas
                # TABLA 1: Columnas A-J (1-10) - Coordinación, Código recuperador, Nombre recuperador, y métricas
                encabezados_tabla1 = [
                    'Coordinación', 'Código\nrecuperador', 'Nombre\nrecuperador',
                    'Cantidad\nPrestada', 'Saldo\nCapital', 'Saldo\nVencido',
                    'Saldo\nTotal', 'Saldo\nRiesgo Capital', 'Saldo\nRiesgo Total', '% MORA'
                ]
                for col_idx, encabezado in enumerate(encabezados_tabla1, start=1):
                    cell = ws_x_recup.cell(row=6, column=col_idx)
                    cell.value = encabezado
                    cell.font = Font(bold=True, size=11, color=color_texto_azul_fuerte)
                    cell.fill = PatternFill(start_color=color_fondo_azul_claro, end_color=color_fondo_azul_claro, fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
                
                # Columna K (11): Vacía - separador entre tablas
                
                # TABLA 2: Columnas L-U (12-21) - Copiar encabezados de la fila 10 (donde pandas los escribió) a la fila 6
                for col_idx in range(1, ws_x_recup.max_column + 1):
                    cell_fila10 = ws_x_recup.cell(row=10, column=col_idx)
                    if cell_fila10.value is not None:
                        # Si es del segundo segmento (columna 12 en adelante), copiar a la fila 6
                        if col_idx >= 12:
                            cell_fila6 = ws_x_recup.cell(row=6, column=col_idx)
                            cell_fila6.value = cell_fila10.value
                            cell_fila6.font = Font(bold=True, size=11, color=color_texto_azul_fuerte)
                            cell_fila6.fill = PatternFill(start_color=color_fondo_azul_claro, end_color=color_fondo_azul_claro, fill_type='solid')
                            cell_fila6.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell_fila6.border = Border(
                                left=Side(style='thin', color='000000'),
                                right=Side(style='thin', color='000000'),
                                top=Side(style='thin', color='000000'),
                                bottom=Side(style='thin', color='000000')
                            )
                        # Limpiar la celda de la fila 10 (solo si no está fusionada)
                        # Limpiar la celda de la fila 10 (solo si no está fusionada)
                        limpiar_celda_segura(cell_fila10)
                
                # Aplicar formato a datos (fila 11+)
                # Formato de moneda a columnas numéricas
                mapeo_rangos_recup = {
                    '0': '0',
                    '1-7 días': '1-7 días',
                    '8-15 días': '8-15 días',
                    '16-30 días': '16-30 días',
                    '31-60 días': '31-60 días',
                    '61-90 días': '61-90 días',
                    'Mayor_90': 'Mayor_90'
                }
                columnas_moneda = ['Cantidad Prestada', 'Saldo capital', 'Saldo vencido', 'Saldo total',
                                  'Saldo riesgo capital', 'Saldo riesgo total'] + list(mapeo_rangos_recup.values())
                for col_name in columnas_moneda:
                    if col_name in df_x_recuperador_ordenado.columns:
                        col_idx = df_x_recuperador_ordenado.columns.get_loc(col_name) + 1
                        for row in range(11, ws_x_recup.max_row + 1):
                            cell = ws_x_recup.cell(row=row, column=col_idx)
                            if cell.value is not None:
                                cell.number_format = EXCEL_CONFIG['currency_format']
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                                cell.border = Border(
                                    left=Side(style='thin', color='000000'),
                                    right=Side(style='thin', color='000000'),
                                    top=Side(style='thin', color='000000'),
                                    bottom=Side(style='thin', color='000000')
                                )
                
                # Formato de porcentaje a % MORA - Solo encabezado en amarillo
                color_amarillo = 'FFFF00'
                if '% MORA' in df_x_recuperador_ordenado.columns:
                    col_idx = df_x_recuperador_ordenado.columns.get_loc('% MORA') + 1
                    cell_header = ws_x_recup.cell(row=6, column=col_idx)
                    cell_header.fill = PatternFill(start_color=color_amarillo, end_color=color_amarillo, fill_type='solid')
                    for row in range(11, ws_x_recup.max_row + 1):
                        cell = ws_x_recup.cell(row=row, column=col_idx)
                        if cell.value is not None:
                            cell.number_format = '0.00%'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            cell.border = Border(
                                left=Side(style='thin', color='000000'),
                                right=Side(style='thin', color='000000'),
                                top=Side(style='thin', color='000000'),
                                bottom=Side(style='thin', color='000000')
                            )
                
                # Formato a columnas de texto (Coordinación, Código recuperador, Nombre recuperador)
                for col_name in ['Coordinación', 'Código recuperador', 'Nombre recuperador']:
                    if col_name in df_x_recuperador_ordenado.columns:
                        col_idx = df_x_recuperador_ordenado.columns.get_loc(col_name) + 1
                        for row in range(11, ws_x_recup.max_row + 1):
                            cell = ws_x_recup.cell(row=row, column=col_idx)
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                            cell.border = Border(
                                left=Side(style='thin', color='000000'),
                                right=Side(style='thin', color='000000'),
                                top=Side(style='thin', color='000000'),
                                bottom=Side(style='thin', color='000000')
                            )
                            # Resaltar fila de Total con azul claro
                            if cell.value == 'Total':
                                for col in range(1, ws_x_recup.max_column + 1):
                                    total_cell = ws_x_recup.cell(row=row, column=col)
                                    total_cell.fill = PatternFill(start_color=color_fondo_azul_claro, end_color=color_fondo_azul_claro, fill_type='solid')
                                    total_cell.font = Font(bold=True, color=color_texto_azul_fuerte)
                                    total_cell.border = Border(
                                        left=Side(style='thin', color='000000'),
                                        right=Side(style='thin', color='000000'),
                                        top=Side(style='thin', color='000000'),
                                        bottom=Side(style='thin', color='000000')
                                    )
                
                # Ajustar ancho de columnas
                for col_idx in range(1, ws_x_recup.max_column + 1):
                    column_letter = get_column_letter(col_idx)
                    if col_idx != 11:  # Columna K (11) será comprimida
                        max_length = 0
                        for row in range(1, ws_x_recup.max_row + 1):
                            cell = ws_x_recup.cell(row=row, column=col_idx)
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        ws_x_recup.column_dimensions[column_letter].width = min(max_length + 2, 20)
                
                # Columna K (11): Completamente oculta (ancho 0 y hidden)
                column_letter_k = get_column_letter(11)
                ws_x_recup.column_dimensions[column_letter_k].width = 0.0
                ws_x_recup.column_dimensions[column_letter_k].hidden = True
                for row in range(1, ws_x_recup.max_row + 1):
                    cell = ws_x_recup.cell(row=row, column=11)
                    limpiar_celda_segura(cell)
                
                # Ajustar altura de filas - Fila 6 visible y larga, filas 7-10 con altura 0 (completamente invisibles)
                ws_x_recup.row_dimensions[6].height = 30
                ws_x_recup.row_dimensions[7].height = 0.0
                ws_x_recup.row_dimensions[8].height = 0.0
                ws_x_recup.row_dimensions[9].height = 0.0
                ws_x_recup.row_dimensions[10].height = 0.0
                
                # Ocultar filas 7, 8, 9 y 10 completamente
                ws_x_recup.row_dimensions[7].hidden = True
                ws_x_recup.row_dimensions[8].hidden = True
                ws_x_recup.row_dimensions[9].hidden = True
                ws_x_recup.row_dimensions[10].hidden = True
                
                # Determinar límites reales de la tabla
                ultima_col_con_datos = 0
                ultima_fila_con_datos = 0
                for row in range(11, ws_x_recup.max_row + 1):
                    for col in range(1, ws_x_recup.max_column + 1):
                        cell = ws_x_recup.cell(row=row, column=col)
                        if cell.value is not None and str(cell.value).strip() != '':
                            ultima_col_con_datos = max(ultima_col_con_datos, col)
                            ultima_fila_con_datos = max(ultima_fila_con_datos, row)
                
                if ultima_col_con_datos == 0:
                    ultima_col_con_datos = 21  # Columna U
                
                # Limpiar celdas fuera del área principal (igual que X_Coordinación)
                for row in range(1, ws_x_recup.max_row + 1):
                    for col in range(1, ws_x_recup.max_column + 1):
                        en_area_principal = False
                        
                        if row == 5:
                            if col >= 10 and col <= 15:  # J-O
                                en_area_principal = True
                        elif row == 6:
                            if (col >= 1 and col <= 10) or (col >= 12 and col <= 21):  # A-J o L-U
                                en_area_principal = True
                        elif row >= 11 and row <= ultima_fila_con_datos:
                            if (col >= 1 and col <= 10) or (col >= 12 and col <= ultima_col_con_datos):
                                en_area_principal = True
                        
                        if not en_area_principal:
                            cell = ws_x_recup.cell(row=row, column=col)
                            if limpiar_celda_segura(cell):
                                try:
                                    cell.font = Font()
                                    cell.alignment = Alignment()
                                except:
                                    pass
                
                logger.info("✅ Hoja 'X_Recuperador' creada exitosamente como SEGUNDA HOJA")
            else:
                logger.warning("⚠️ No se pudo crear la hoja 'X_Recuperador' (DataFrame vacío)")
            
            # --- Hoja 1: Informe completo ---
            hoja_informe = fecha_actual
            
            
            # DIAGNÓSTICO FINAL: Verificar columnas antes de escribir
            columnas_finales = [col for col in df_completo_sin_links.columns if 'par' in str(col).lower()]
            if columnas_finales:
                logger.error(f"🚨 ERROR CRÍTICO: Columnas PAR en Informe Completo FINAL: {columnas_finales}")
            else:
                logger.info(f"✅ Informe Completo FINAL sin columnas PAR")
            
            # Agregar columna 'Concepto Depósito' al informe completo
            df_completo_sin_links = agregar_columna_concepto_deposito(df_completo_sin_links.copy())
            
            # Agregar columnas 'Saldo riesgo capital', 'Saldo riesgo total' y '% MORA'
            df_completo_sin_links = agregar_columnas_riesgo_y_mora(df_completo_sin_links.copy())
            df_completo_sin_links = agregar_columnas_dias_ultimo_pago_y_alerta(df_completo_sin_links)

            df_completo_sin_links.to_excel(writer, sheet_name=hoja_informe, index=False, startrow=1)
            ws_informe = writer.sheets[hoja_informe]
            
            # *** NUEVO: Aplicar formato de texto a 'Código acreditado' y 'Concepto Depósito' para preservar ceros ***
            if 'Código acreditado' in df_completo_sin_links.columns:
                # Buscar la columna 'Código acreditado' en el Excel
                for col_idx in range(1, ws_informe.max_column + 1):
                    if ws_informe.cell(row=2, column=col_idx).value == 'Código acreditado':
                        # Aplicar formato de texto desde fila 3 (datos) hasta el final
                        for row in range(3, ws_informe.max_row + 1):
                            ws_informe.cell(row=row, column=col_idx).number_format = '@'
                        logger.info(f"✅ Formato de texto aplicado a columna 'Código acreditado' (columna {col_idx})")
                        break
            
            # Aplicar formato de texto a 'Concepto Depósito'
            aplicar_formato_texto_concepto_deposito(ws_informe, df_completo_sin_links)
            
            # Aplicar formato condicional a la hoja de informe completo
            aplicar_formato_condicional(ws_informe, columna_mora, len(df_completo))
            
            # Añadir hipervínculos si existe la columna 'Link de Geolocalización'
            if 'Link de Geolocalización' in df_completo_sin_links.columns:
                link_col = df_completo_sin_links.columns.get_loc('Link de Geolocalización') + 1  # +1 porque Excel es 1-indexado
                
                # Escribir hipervínculos usando fórmula HYPERLINK (persisten mejor con openpyxl)
                for i, (idx, row) in enumerate(df_completo.iterrows()):
                    row_num = i + 3  # +3 porque Excel empieza en 1, hay títulos en fila 1, encabezados en fila 2, datos empiezan en fila 3
                    if 'link_texto' in df_completo.columns and 'link_url' in df_completo.columns:
                        texto = row['link_texto']
                        url = row['link_url']
                        escribir_hipervinculo_excel(ws_informe, row_num, link_col, texto, url)
            
            # Aplicar formato final y formatos específicos antes de crear la tabla
            aplicar_formato_final(ws_informe, df_completo_sin_links, es_hoja_mora=False)
            aplicar_formato_porcentaje_mora(ws_informe, df_completo_sin_links)
            aplicar_formato_alerta(ws_informe, df_completo_sin_links)
            # Crear tabla al final para no afectar hipervínculos ni formato condicional
            crear_tabla_excel(ws_informe, df_completo_sin_links, hoja_informe, incluir_columnas_adicionales=False)

            # --- Hoja RECUPERADOR_000124 (registros con código recuperador en CODIGOS_RECUPERADOR_EXCLUIR) ---
            if df_recup_000124_sin_links is not None and len(df_recup_000124_sin_links) > 0:
                df_recup_000124_sin_links.to_excel(writer, sheet_name='RECUPERADOR_000124', index=False, startrow=1)
                ws_recup = writer.sheets['RECUPERADOR_000124']
                if 'Código acreditado' in df_recup_000124_sin_links.columns:
                    for col_idx in range(1, ws_recup.max_column + 1):
                        if ws_recup.cell(row=2, column=col_idx).value == 'Código acreditado':
                            for row in range(3, ws_recup.max_row + 1):
                                ws_recup.cell(row=row, column=col_idx).number_format = '@'
                            break
                aplicar_formato_texto_concepto_deposito(ws_recup, df_recup_000124_sin_links)
                aplicar_formato_condicional(ws_recup, columna_mora, len(df_recup_000124_sin_links))
                if 'Link de Geolocalización' in df_recup_000124_sin_links.columns and df_recup_000124_completo is not None:
                    link_col_recup = df_recup_000124_sin_links.columns.get_loc('Link de Geolocalización') + 1
                    for i, (_, row) in enumerate(df_recup_000124_completo.iterrows()):
                        row_num = i + 3
                        if 'link_texto' in df_recup_000124_completo.columns and 'link_url' in df_recup_000124_completo.columns:
                            escribir_hipervinculo_excel(ws_recup, row_num, link_col_recup, row['link_texto'], row['link_url'])
                aplicar_formato_final(ws_recup, df_recup_000124_sin_links, es_hoja_mora=False)
                aplicar_formato_porcentaje_mora(ws_recup, df_recup_000124_sin_links)
                aplicar_formato_alerta(ws_recup, df_recup_000124_sin_links)
                crear_tabla_excel(ws_recup, df_recup_000124_sin_links, 'RECUPERADOR_000124', incluir_columnas_adicionales=False)
                logger.info(f"✅ Hoja RECUPERADOR_000124 creada con {len(df_recup_000124_sin_links)} registros")

            # --- PASO 6.1: Crear hoja "Mora" ---
            # Verificar columnas duplicadas antes de escribir hoja Mora
            columnas_duplicadas_mora = df_mora.columns[df_mora.columns.duplicated()].tolist()
            if columnas_duplicadas_mora:
                logger.error(f"🚨 COLUMNAS DUPLICADAS encontradas en df_mora: {columnas_duplicadas_mora}")
                # Eliminar columnas duplicadas
                df_mora = df_mora.loc[:, ~df_mora.columns.duplicated()]
                logger.info(f"🗑️ Columnas duplicadas eliminadas de df_mora")
            
            # DIAGNÓSTICO FINAL: Verificar columnas PAR en df_mora
            columnas_par_mora = [col for col in df_mora.columns if 'par' in str(col).lower()]
            if columnas_par_mora:
                logger.error(f"🚨 ERROR CRÍTICO: Columnas PAR en Mora FINAL: {columnas_par_mora}")
            else:
                logger.info(f"✅ Mora FINAL sin columnas PAR")
            
            # Crear DataFrame sin las columnas temporales de links para escritura en Excel
            df_mora_sin_links = df_mora.drop(columns=['link_texto', 'link_url'], errors='ignore')
            
            # Agregar columna 'Concepto Depósito' a la hoja Mora
            df_mora_sin_links = agregar_columna_concepto_deposito(df_mora_sin_links.copy())
            
            # Agregar columnas 'Saldo riesgo capital', 'Saldo riesgo total' y '% MORA'
            df_mora_sin_links = agregar_columnas_riesgo_y_mora(df_mora_sin_links.copy())
            
            df_mora_sin_links.to_excel(writer, sheet_name='Mora', index=False, startrow=1)
            
            # Aplicar formato condicional
            worksheet_mora = writer.sheets['Mora']
            aplicar_formato_condicional(worksheet_mora, columna_mora, len(df_mora))
            
            # Añadir hipervínculos si existe la columna 'Link de Geolocalización'
            if 'Link de Geolocalización' in df_mora_sin_links.columns:
                link_col = df_mora_sin_links.columns.get_loc('Link de Geolocalización') + 1  # +1 porque Excel es 1-indexado
                
                # Escribir hipervínculos usando los datos originales de df_mora
                for i, (idx, row) in enumerate(df_mora.iterrows()):
                    row_num = i + 3  # +3 porque Excel empieza en 1, hay títulos en fila 1, encabezados en fila 2, datos empiezan en fila 3
                    if 'link_texto' in df_mora.columns and 'link_url' in df_mora.columns:
                        texto = row['link_texto']
                        url = row['link_url']
                        escribir_hipervinculo_excel(worksheet_mora, row_num, link_col, texto, url)
            
            # Aplicar formato de texto a 'Concepto Depósito'
            aplicar_formato_texto_concepto_deposito(worksheet_mora, df_mora_sin_links)
            
            # Crear tabla formal de Excel para la hoja Mora y formato final
            crear_tabla_excel(worksheet_mora, df_mora_sin_links, 'Mora', incluir_columnas_adicionales=True)
            aplicar_formato_final(worksheet_mora, df_mora_sin_links, es_hoja_mora=True)
            aplicar_formato_porcentaje_mora(worksheet_mora, df_mora_sin_links)

            # --- PASO 6.1.1: Crear hoja "Cuentas con saldo vencido" ---
            if df_saldo_vencido is not None and len(df_saldo_vencido) > 0:
                # Verificar columnas duplicadas antes de escribir hoja Saldo Vencido
                columnas_duplicadas_saldo = df_saldo_vencido.columns[df_saldo_vencido.columns.duplicated()].tolist()
                if columnas_duplicadas_saldo:
                    logger.error(f"🚨 COLUMNAS DUPLICADAS encontradas en df_saldo_vencido: {columnas_duplicadas_saldo}")
                    # Eliminar columnas duplicadas
                    df_saldo_vencido = df_saldo_vencido.loc[:, ~df_saldo_vencido.columns.duplicated()]
                    logger.info(f"🗑️ Columnas duplicadas eliminadas de df_saldo_vencido")
                
                # DIAGNÓSTICO FINAL: Verificar columnas PAR en df_saldo_vencido
                columnas_par_saldo = [col for col in df_saldo_vencido.columns if 'par' in str(col).lower()]
                if columnas_par_saldo:
                    logger.error(f"🚨 ERROR CRÍTICO: Columnas PAR en Saldo Vencido FINAL: {columnas_par_saldo}")
                else:
                    logger.info(f"✅ Saldo Vencido FINAL sin columnas PAR")
                
                # Crear DataFrame sin las columnas temporales de links para escritura en Excel
                df_saldo_vencido_sin_links = df_saldo_vencido.drop(columns=['link_texto', 'link_url'], errors='ignore')
                
                # Agregar columna 'Concepto Depósito' a la hoja Cuentas con saldo vencido
                df_saldo_vencido_sin_links = agregar_columna_concepto_deposito(df_saldo_vencido_sin_links.copy())
                
                # Agregar columnas 'Saldo riesgo capital', 'Saldo riesgo total' y '% MORA'
                df_saldo_vencido_sin_links = agregar_columnas_riesgo_y_mora(df_saldo_vencido_sin_links.copy())
                
                df_saldo_vencido_sin_links.to_excel(writer, sheet_name='Cuentas con saldo vencido', index=False, startrow=1)
                
                # NO aplicar formato condicional para la hoja "Cuentas con saldo vencido"
                worksheet_saldo = writer.sheets['Cuentas con saldo vencido']
                # aplicar_formato_condicional(worksheet_saldo, columna_mora, len(df_saldo_vencido))  # Comentado: no queremos colores en esta hoja
                
                # Aplicar formato de texto a 'Concepto Depósito'
                aplicar_formato_texto_concepto_deposito(worksheet_saldo, df_saldo_vencido_sin_links)
                
                # Añadir hipervínculos si existe la columna 'Link de Geolocalización'
                if 'Link de Geolocalización' in df_saldo_vencido_sin_links.columns:
                    link_col = df_saldo_vencido_sin_links.columns.get_loc('Link de Geolocalización') + 1  # +1 porque Excel es 1-indexado
                    
                    # Escribir hipervínculos usando los datos originales de df_saldo_vencido
                    for i, (idx, row) in enumerate(df_saldo_vencido.iterrows()):
                        row_num = i + 3  # +3 porque Excel empieza en 1, hay títulos en fila 1, encabezados en fila 2, datos empiezan en fila 3
                        if 'link_texto' in df_saldo_vencido.columns and 'link_url' in df_saldo_vencido.columns:
                            texto = row['link_texto']
                            url = row['link_url']
                            escribir_hipervinculo_excel(worksheet_saldo, row_num, link_col, texto, url)
                
                # Crear tabla formal de Excel para la hoja Saldo Vencido y formato final
                crear_tabla_excel(worksheet_saldo, df_saldo_vencido_sin_links, 'Cuentas con saldo vencido', incluir_columnas_adicionales=False)
                aplicar_formato_final(worksheet_saldo, df_saldo_vencido_sin_links, es_hoja_mora=False)
                aplicar_formato_porcentaje_mora(worksheet_saldo, df_saldo_vencido_sin_links)
                
                logger.info(f"✅ Hoja 'Cuentas con saldo vencido' creada con {len(df_saldo_vencido)} registros")
            else:
                logger.info("⚠️ No se creó la hoja 'Cuentas con saldo vencido' (no hay datos o columna faltante)")

            # --- PASO 6.1.2: Crear hoja "Liquidación anticipada" ---
            logger.info("Creando hoja 'Liquidación anticipada'")
            
            # Validar columnas requeridas para liquidación anticipada
            # Mapear a los nombres reales de columnas basados en el contenido de la primera fila
            columnas_requeridas = {
                'ciclo': 'Ciclo',
                'nombre_acreditado': 'Nombre acreditado', 
                'intereses_vencidos': 'Saldo interés vencido',
                'comision_vencida': 'Saldo comisión vencida',
                'recargos': 'Saldo recargos',
                'saldo_capital': 'Saldo capital'
            }
            
            # Verificar qué columnas existen en df_completo
            logger.info(f"🔍 DIAGNÓSTICO DETALLADO DE COLUMNAS:")
            logger.info(f"   - Columnas disponibles en df_completo: {list(df_completo.columns)}")
            logger.info(f"   - Columnas requeridas: {list(columnas_requeridas.values())}")
            
            # Función para buscar columnas por similitud
            def buscar_columna_similar(columna_requerida, columnas_disponibles):
                """Busca una columna por similitud de nombre"""
                columna_requerida_clean = columna_requerida.lower().replace(' ', '').replace('ó', 'o').replace('í', 'i').replace('é', 'e').replace('á', 'a').replace('ú', 'u')
                
                logger.info(f"🔍 Buscando columna similar a '{columna_requerida}' (limpio: '{columna_requerida_clean}')")
                
                for col_disponible in columnas_disponibles:
                    col_disponible_clean = col_disponible.lower().replace(' ', '').replace('ó', 'o').replace('í', 'i').replace('é', 'e').replace('á', 'a').replace('ú', 'u')
                    
                    # Búsqueda exacta
                    if col_disponible_clean == columna_requerida_clean:
                        logger.info(f"✅ Encontrada coincidencia exacta: '{col_disponible}'")
                        return col_disponible
                    
                    # Búsqueda por coincidencia parcial
                    if columna_requerida_clean in col_disponible_clean or col_disponible_clean in columna_requerida_clean:
                        logger.info(f"✅ Encontrada coincidencia parcial: '{col_disponible}'")
                        return col_disponible
                
                logger.warning(f"❌ No se encontró columna similar a '{columna_requerida}'")
                return None
            
            # Mapear columnas requeridas a columnas reales
            columnas_mapeadas = {}
            columnas_faltantes = []
            
            # Mapeo manual basado en la estructura real del archivo
            # Según el debug anterior, las columnas están en posiciones específicas
            mapeo_manual = {
                'ciclo': df_completo.columns[7] if len(df_completo.columns) > 7 else None,  # Unnamed: 7 (Columna 8 Excel)
                'nombre_acreditado': df_completo.columns[8] if len(df_completo.columns) > 8 else None,  # Unnamed: 8 (Columna 9 Excel)
                'intereses_vencidos': df_completo.columns[25] if len(df_completo.columns) > 25 else None,  # Unnamed: 25 (Columna 26 Excel - Saldo interés vencido)
                'comision_vencida': df_completo.columns[26] if len(df_completo.columns) > 26 else None,  # Unnamed: 26 (Columna 27 Excel - Saldo comisión vencida)
                'recargos': df_completo.columns[27] if len(df_completo.columns) > 27 else None,  # Unnamed: 27 (Columna 28 Excel - Saldo recargos)
                'saldo_capital': df_completo.columns[22] if len(df_completo.columns) > 22 else None,  # Unnamed: 22 (Columna 23 Excel - Saldo capital)
            }
            
            for key, columna_requerida in columnas_requeridas.items():
                # Primero intentar mapeo manual
                columna_manual = mapeo_manual.get(key)
                if columna_manual:
                    columnas_mapeadas[key] = columna_manual
                    logger.info(f"✅ Columna '{columna_requerida}' mapeada manualmente a '{columna_manual}'")
                elif columna_requerida in df_completo.columns:
                    columnas_mapeadas[key] = columna_requerida
                    logger.info(f"✅ Columna '{columna_requerida}' encontrada exactamente")
                else:
                    # Buscar por similitud
                    columna_encontrada = buscar_columna_similar(columna_requerida, df_completo.columns)
                    if columna_encontrada:
                        columnas_mapeadas[key] = columna_encontrada
                        logger.info(f"✅ Columna '{columna_requerida}' mapeada por similitud a '{columna_encontrada}'")
                    else:
                        columnas_faltantes.append(columna_requerida)
                        logger.warning(f"⚠️ Columna '{columna_requerida}' no encontrada para 'Liquidación anticipada'")
            
            if not columnas_faltantes:
                logger.info("✅ Todas las columnas requeridas para liquidación anticipada están disponibles")
            else:
                logger.warning(f"⚠️ Columnas faltantes: {columnas_faltantes}")
            
            # Definir columnas para la hoja de liquidación anticipada
            liquidacion_columns = [
                COLUMN_MAPPING.get('codigo', 'Código acreditado'),  # A
                'Ciclo',                                           # B
                'Nombre del acreditado',                           # C
                'Saldo interés vencido',                           # D
                'Saldo comisión vencida',                          # E
                'Saldo recargos',                                  # F
                'Saldo capital',                                   # G
                'Saldo adelantado',                                # H (NUEVA COLUMNA)
                'Intereses del próximo pago sin vencer',           # I
                'Comisiones del próximo pago sin vencer',          # J
                'Cantidad a liquidar',                             # K
                'Cálculo válido hasta el próximo pago'             # L
            ]
            
            # Crear DataFrame vacío para la hoja de liquidación anticipada
            df_liquidacion = pd.DataFrame(columns=liquidacion_columns)
            
            # Inicializar fila con datos vacíos
            fila_inicial = [''] * len(liquidacion_columns)
            df_liquidacion.loc[0] = fila_inicial
            
            # Escribir la hoja
            df_liquidacion.to_excel(writer, sheet_name='Liquidación anticipada', index=False, startrow=1)
            ws_liquidacion = writer.sheets['Liquidación anticipada']
            
            # --- Diseño personalizado para la hoja de liquidación anticipada ---
            
            # 1. Combinar celdas D1:F1 para "Montos Vencidos" con relleno azul claro
            ws_liquidacion.merge_cells('D1:F1')
            celda_titulo = ws_liquidacion['D1']
            celda_titulo.value = 'Montos Vencidos'
            celda_titulo.fill = PatternFill(start_color=COLORS['light_blue'], end_color=COLORS['light_blue'], fill_type='solid')
            celda_titulo.font = Font(bold=True)
            celda_titulo.alignment = Alignment(horizontal='center', vertical='center')
            
            # 2. Establecer ancho de columna optimizado para cada tipo de dato
            widths = {
                'A': 18,  # Código acreditado (más ancho para códigos largos)
                'B': 12,  # Ciclo
                'C': 25,  # Nombre (más ancho para nombres largos)
                'D': 20,  # Saldo interés vencido
                'E': 20,  # Saldo comisión vencida
                'F': 15,  # Saldo recargos
                'G': 18,  # Saldo capital
                'H': 18,  # Saldo adelantado (NUEVA COLUMNA)
                'I': 22,  # Intereses próximo pago
                'J': 22,  # Comisiones próximo pago
                'K': 20,  # Cantidad a liquidar
                'L': 25   # Cálculo válido hasta
            }
            
            for col_letter, width in widths.items():
                ws_liquidacion.column_dimensions[col_letter].width = width
            
            # 3. Establecer altura de filas para mejor legibilidad
            ws_liquidacion.row_dimensions[2].height = 35  # Encabezados más altos
            ws_liquidacion.row_dimensions[3].height = 30  # Datos más altos
            
            # 4. Aplicar formato minimalista pero legible a todas las celdas
            # Encabezados (fila 2)
            for col in range(1, 13):  # Columnas A a L
                cell = ws_liquidacion.cell(row=2, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(name='Arial', size=10, bold=True, color='2F4F4F')  # Azul gris oscuro
                cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')  # Gris muy claro
                cell.border = Border(
                    left=Side(style='thin', color='D3D3D3'),
                    right=Side(style='thin', color='D3D3D3'),
                    top=Side(style='thin', color='D3D3D3'),
                    bottom=Side(style='thin', color='D3D3D3')
                )
            
            # Datos (fila 3)
            for col in range(1, 13):  # Columnas A a L
                cell = ws_liquidacion.cell(row=3, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(name='Arial', size=10, color='2C2C2C')  # Gris oscuro
                cell.border = Border(
                    left=Side(style='thin', color='D3D3D3'),
                    right=Side(style='thin', color='D3D3D3'),
                    top=Side(style='thin', color='D3D3D3'),
                    bottom=Side(style='thin', color='D3D3D3')
                )
            
            # 4. Aplicar relleno verde claro a celdas manuales (columnas I, J, L) con formato mejorado
            celdas_manuales = ['I3', 'J3', 'L3']
            for celda in celdas_manuales:
                ws_liquidacion[celda].fill = PatternFill(start_color=COLORS['light_green'], end_color=COLORS['light_green'], fill_type='solid')
                ws_liquidacion[celda].font = Font(name='Arial', size=10, bold=True, color='2C2C2C')  # Texto en negrita para celdas manuales
            
            # 5. Aplicar formato de moneda a columnas D:K (4:11)
            for col_letter in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                ws_liquidacion[f'{col_letter}3'].number_format = EXCEL_CONFIG['currency_format']
            
            # 5. Formatear celda A3 (Código acreditado) para mantener ceros al inicio
            ws_liquidacion['A3'].number_format = '@'  # Formato de texto para mantener ceros
            
            # 6. Agregar fórmulas BUSCARV en B3:G3 para autocompletado desde "Informe Completo"
            # Obtener el nombre de la hoja "Informe Completo" (fecha_actual)
            nombre_hoja_informe = hoja_informe  # Ya definido anteriormente
            
            # *** CORREGIDO: Usar columnas completas en lugar de rango fijo ***
            ultima_columna_informe = len(df_completo.columns)
            ultima_col_letter = get_column_letter(ultima_columna_informe)
            rango_informe = f"$A:{ultima_col_letter}"
            
            logger.info(f"🔍 Debug fórmulas BUSCARV:")
            logger.info(f"   - Nombre hoja informe: '{nombre_hoja_informe}'")
            logger.info(f"   - Rango informe: '{rango_informe}'")
            logger.info(f"   - Registros en df_completo: {len(df_completo)}")
            logger.info(f"   - Columnas en df_completo: {list(df_completo.columns)}")
            
            # Mostrar las primeras filas para verificar datos
            if len(df_completo) > 0:
                primera_columna = df_completo.columns[0]  # Primera columna (debería ser 'Código acreditado')
                logger.info(f"   - Primera columna: '{primera_columna}'")
                logger.info(f"   - Primeros 3 valores de '{primera_columna}': {df_completo[primera_columna].head(3).tolist()}")
            else:
                logger.error("   - ERROR: df_completo está vacío!")
            
            # B3: Ciclo - con manejo de valores nulos usando VLOOKUP (inglés)
            if 'ciclo' in columnas_mapeadas:
                col_ciclo_index = df_completo.columns.get_loc(columnas_mapeadas['ciclo']) + 1  # +1 porque Excel es 1-indexado
                formula_ciclo = f"=IFERROR(VLOOKUP(A3,'{nombre_hoja_informe}'!{rango_informe},{col_ciclo_index},FALSE),\"\")"
                ws_liquidacion['B3'] = formula_ciclo
                logger.info(f"✅ Fórmula B3 (Ciclo): {formula_ciclo}")
            else:
                ws_liquidacion['B3'] = '""'
                logger.warning(f"⚠️ Columna 'Ciclo' no encontrada, B3 quedará vacío")
            
            # C3: Nombre del acreditado - con manejo de valores nulos usando VLOOKUP (inglés)
            if 'nombre_acreditado' in columnas_mapeadas:
                col_nombre_index = df_completo.columns.get_loc(columnas_mapeadas['nombre_acreditado']) + 1
                formula_nombre = f"=IFERROR(VLOOKUP(A3,'{nombre_hoja_informe}'!{rango_informe},{col_nombre_index},FALSE),\"\")"
                ws_liquidacion['C3'] = formula_nombre
                logger.info(f"✅ Fórmula C3 (Nombre): {formula_nombre}")
            else:
                ws_liquidacion['C3'] = '""'
                logger.warning(f"⚠️ Columna 'Nombre acreditado' no encontrada, C3 quedará vacío")
            
            # D3: Saldo interés vencido
            if 'intereses_vencidos' in columnas_mapeadas:
                col_intereses_index = df_completo.columns.get_loc(columnas_mapeadas['intereses_vencidos']) + 1
                formula_intereses = f"=IFERROR(VLOOKUP(A3,'{nombre_hoja_informe}'!{rango_informe},{col_intereses_index},FALSE),0)"
                ws_liquidacion['D3'] = formula_intereses
                logger.info(f"✅ Fórmula D3 (Intereses): {formula_intereses}")
            else:
                ws_liquidacion['D3'] = '0'
                logger.warning(f"⚠️ Columna 'Intereses vencidos' no encontrada, D3 = 0")
            
            # E3: Saldo comisión vencida
            if 'comision_vencida' in columnas_mapeadas:
                col_comision_index = df_completo.columns.get_loc(columnas_mapeadas['comision_vencida']) + 1
                formula_comision = f"=IFERROR(VLOOKUP(A3,'{nombre_hoja_informe}'!{rango_informe},{col_comision_index},FALSE),0)"
                ws_liquidacion['E3'] = formula_comision
                logger.info(f"✅ Fórmula E3 (Comisión): {formula_comision}")
            else:
                ws_liquidacion['E3'] = '0'
                logger.warning(f"⚠️ Columna 'Comisión vencida' no encontrada, E3 = 0")
            
            # F3: Saldo recargos
            if 'recargos' in columnas_mapeadas:
                col_recargos_index = df_completo.columns.get_loc(columnas_mapeadas['recargos']) + 1
                formula_recargos = f"=IFERROR(VLOOKUP(A3,'{nombre_hoja_informe}'!{rango_informe},{col_recargos_index},FALSE),0)"
                ws_liquidacion['F3'] = formula_recargos
                logger.info(f"✅ Fórmula F3 (Recargos): {formula_recargos}")
            else:
                ws_liquidacion['F3'] = '0'
                logger.warning(f"⚠️ Columna 'Recargos' no encontrada, F3 = 0")
            
            # G3: Saldo capital
            if 'saldo_capital' in columnas_mapeadas:
                col_capital_index = df_completo.columns.get_loc(columnas_mapeadas['saldo_capital']) + 1
                formula_capital = f"=IFERROR(VLOOKUP(A3,'{nombre_hoja_informe}'!{rango_informe},{col_capital_index},FALSE),0)"
                ws_liquidacion['G3'] = formula_capital
                logger.info(f"✅ Fórmula G3 (Capital): {formula_capital}")
            else:
                ws_liquidacion['G3'] = '0'
                logger.warning(f"⚠️ Columna 'Saldo capital' no encontrada, G3 = 0")
            
            # H3: Saldo adelantado (NUEVA COLUMNA) - Buscar en columna específica del informe
            # Esta columna busca directamente en una posición fija según el archivo del jefe
            formula_adelantado = f"=IFERROR(VLOOKUP(A3,'{nombre_hoja_informe}'!{rango_informe},52,FALSE),0)"
            ws_liquidacion['H3'] = formula_adelantado
            logger.info(f"✅ Fórmula H3 (Saldo adelantado): {formula_adelantado}")
            
            # 7. Establecer fórmula de suma en columna K3 para "Cantidad a liquidar"
            # Suma: Saldo interés vencido + Saldo comisión + Saldo recargos + Saldo capital + Intereses próximo pago + Comisiones próximo pago
            # Resta: Saldo adelantado
            ws_liquidacion['K3'] = '=SUM(D3:G3,I3:J3)-H3'
            
            # 8. NO usar crear_tabla_excel para mantener el diseño personalizado minimalista
            
            # 9. Aplicar formato final personalizado sin tabla formal
            # Inmovilizar paneles en A3 para mejor navegación
            ws_liquidacion.freeze_panes = 'A3'
            
            # 10. Asegurar que las celdas fuera del área principal tengan fondo blanco
            for row in range(1, 15):  # Filas 1-14
                for col in range(1, 15):  # Columnas A-N
                    cell = ws_liquidacion.cell(row=row, column=col)
                    # Solo aplicar relleno blanco si no tiene relleno especial
                    if cell.fill.start_color.index == '00000000':  # Sin relleno
                        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            # 11. Instrucciones removidas para diseño más limpio
            
            logger.info("✅ Hoja 'Liquidación anticipada' creada")

            # --- PASO 6.2: Crear hojas por coordinación --- [ELIMINADO - iteración 1]
            # Las hojas por coordinación (Atlacomulco, Maravatio, Metepec, etc.) fueron eliminadas
            # del nuevo diseño. El desglose por coordinación ahora vive en los pivots de X_Coordinación.
            pass

        logger.info(f"Procesamiento completado exitosamente. Archivo generado: {ruta_salida}")
        
        return ruta_salida, len(coordinaciones_data)
        
    except FileNotFoundError as e:
        logger.error(f"Archivo no encontrado: {str(e)}")
        raise Exception(f"El archivo especificado no existe: {str(e)}")
    except pd.errors.EmptyDataError as e:
        logger.error(f"Archivo Excel vacío: {str(e)}")
        raise Exception(f"El archivo Excel está vacío o no contiene datos válidos: {str(e)}")
    except (ValueError, OSError, IOError) as e:
        logger.error(f"Error al leer archivo Excel: {str(e)}")
        raise Exception(f"Error al leer el archivo Excel. Verifique que sea un archivo válido: {str(e)}")
    except ValueError as e:
        logger.error(f"Error de validación: {str(e)}")
        raise Exception(f"Error de validación de datos: {str(e)}")
    except Exception as e:
        logger.error(f"Error inesperado procesando archivo: {str(e)}")
        raise Exception(f"Error inesperado procesando archivo: {str(e)}")

@reportes_bp.route('/antiguedad')
def antiguedad_form():
    """Página para subir archivo de reporte de antigüedad"""
    return render_template('antiguedad.html')

@reportes_bp.route('/download/<filename>')
def download_file(filename):
    """Servir archivos de descarga"""
    file_path = os.path.join('static', 'downloads', filename)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    else:
        return "Archivo no encontrado", 404

@reportes_bp.route('/procesar_antiguedad', methods=['POST'])
@login_required
@require_permission('generate_reports')
def procesar_antiguedad():
    """Procesa el archivo subido y devuelve el reporte"""
    if 'archivo' not in request.files:
        return Response('No se seleccionó ningún archivo', status=400)
    
    archivo = request.files['archivo']
    if archivo.filename == '':
        return Response('No se seleccionó ningún archivo', status=400)
    
    if not allowed_file(archivo.filename):
        return Response('El archivo debe ser de tipo Excel (.xlsx o .xls)', status=400)
    
    try:
        # Validar tamaño del archivo
        archivo.seek(0, 2)  # Ir al final del archivo
        file_size = archivo.tell()
        archivo.seek(0)  # Volver al inicio
        
        if file_size > MAX_FILE_SIZE:
            return Response(f'El archivo es demasiado grande. Tamaño máximo permitido: {MAX_FILE_SIZE // (1024*1024)}MB', status=400)
        
        # Guardar archivo temporalmente
        filename = secure_filename(archivo.filename)
        archivo_path = os.path.join(UPLOAD_FOLDER, filename)
        archivo.save(archivo_path)
        
        logger.info(f"Archivo subido exitosamente: {filename} ({file_size} bytes)")
        
        ruta_salida, num_coordinaciones = procesar_reporte_antiguedad(archivo_path, codigos_a_excluir=None)
        
        # Mover al directorio de reportes SIN modificar el nombre
        import shutil
        os.makedirs(REPORTS_FOLDER, exist_ok=True)
        ruta_final = os.path.join(REPORTS_FOLDER, os.path.basename(ruta_salida))
        shutil.move(ruta_salida, ruta_final)
        
        # Guardar en el historial de reportes
        try:
            file_size = os.path.getsize(ruta_final)
            report_history = ReportHistory(
                user_id=current_user.id,
                report_type='individual',
                filename=os.path.basename(ruta_final),
                file_path=ruta_final,  # Usar ruta final en directorio de reportes
                file_size=file_size
            )
            db.session.add(report_history)
            db.session.commit()
            logger.info(f"Reporte principal guardado en historial: {os.path.basename(ruta_final)}")
        except Exception as e:
            logger.error(f"Error guardando reporte principal: {str(e)}")
            db.session.rollback()
        
        # Limpiar archivo temporal original
        try:
            os.remove(archivo_path)
        except (OSError, FileNotFoundError):
            pass
        
        # Devolver el archivo generado directamente
        return send_file(
            ruta_final,  # Usar ruta final en directorio de reportes
            as_attachment=True,
            download_name=os.path.basename(ruta_final),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error en procesamiento de archivo: {str(e)}")
        return Response(f'Error procesando archivo: {str(e)}', status=500)

def detectar_tipo_archivo(df):
    """Detecta el tipo de archivo por su estructura/columnas"""
    columnas_str = ' '.join([str(col).lower() for col in df.columns])
    
    # Detectar "Cobranza.xlsx"
    if 'cobranza' in columnas_str or 'call_center' in columnas_str or 'estatus' in columnas_str:
        return 'cobranza'
    
    # Detectar "Reporte de conformacion de grupo"  
    elif 'conformacion' in columnas_str or 'grupo' in columnas_str:
        return 'conformacion_grupo'
    
    # Detectar "Ahorros"
    elif 'ahorro' in columnas_str or 'deposito' in columnas_str:
        return 'ahorros'
    
    # Detectar "ReportedeAntiguedaddeCarteraGrupal"
    elif 'grupal' in columnas_str and 'antiguedad' in columnas_str:
        return 'antiguedad_grupal'
    
    # Detectar "Situacion de cartera.xls"
    elif 'situacion' in columnas_str or 'estado' in columnas_str:
        return 'situacion_cartera'
    
    else:
        return 'desconocido'

@reportes_bp.route('/procesar_antiguedad_grupal', methods=['POST'])
@login_required
@require_permission('generate_reports')
def procesar_antiguedad_grupal():
    """Procesa los 5 archivos subidos para el reporte grupal"""
    if 'archivos' not in request.files:
        return Response('No se seleccionaron archivos', status=400)
    
    archivos = request.files.getlist('archivos')
    
    if len(archivos) != 5:
        return Response('Deben seleccionarse exactamente 5 archivos', status=400)
    
    # Validar todos los archivos
    for archivo in archivos:
        if archivo.filename == '':
            return Response('Uno o más archivos están vacíos', status=400)
        
        if not allowed_file(archivo.filename):
            return Response(f'El archivo {archivo.filename} debe ser de tipo Excel (.xlsx o .xls)', status=400)
    
    try:
        # Guardar archivos temporalmente y detectar tipos
        archivos_info = []
        archivos_paths = []
        
        for archivo in archivos:
            filename = secure_filename(archivo.filename)
            archivo_path = os.path.join(UPLOAD_FOLDER, filename)
            archivo.save(archivo_path)
            archivos_paths.append(archivo_path)
            
            # Detectar tipo de archivo
            df_temp = pd.read_excel(archivo_path, engine='openpyxl', dtype=DTYPE_CONFIG, header=0)
            tipo = detectar_tipo_archivo(df_temp)
            
            archivos_info.append({
                'filename': filename,
                'path': archivo_path,
                'tipo': tipo,
                'df': df_temp
            })
            
            logger.info(f"Archivo {filename} detectado como tipo: {tipo}")
        
        # Verificar que se detectaron todos los tipos requeridos
        tipos_requeridos = ['cobranza', 'conformacion_grupo', 'ahorros', 'antiguedad_grupal', 'situacion_cartera']
        tipos_detectados = [info['tipo'] for info in archivos_info]
        
        tipos_faltantes = set(tipos_requeridos) - set(tipos_detectados)
        if tipos_faltantes:
            return Response(f'Faltan tipos de archivo: {", ".join(tipos_faltantes)}', status=400)
        
        # Procesar reporte grupal (por ahora, solo devolver un mensaje de éxito)
        logger.info("Iniciando procesamiento de reporte grupal...")
        
        # TODO: Implementar la lógica de consolidación de los 5 archivos
        # Por ahora, solo creamos un archivo de ejemplo
        
        # Limpiar archivos temporales
        for archivo_path in archivos_paths:
            try:
                os.remove(archivo_path)
            except (OSError, FileNotFoundError):
                pass
        
        # Crear un archivo Excel de ejemplo para el reporte grupal
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Grupal"
        ws['A1'] = "Reporte de Antigüedad Grupal"
        ws['A2'] = "Archivos procesados:"
        
        for i, info in enumerate(archivos_info, start=3):
            ws[f'A{i}'] = f"{info['filename']} - Tipo: {info['tipo']}"
        
        # Guardar archivo temporal
        ruta_salida = os.path.join(UPLOAD_FOLDER, f"reporte_grupal_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(ruta_salida)
        
        # *** NUEVO: Mover archivo al directorio de reportes permanentes ***
        ruta_final = move_to_reports_folder(ruta_salida, 'grupal')
        
        # Guardar en el historial de reportes
        try:
            file_size = os.path.getsize(ruta_final)
            report_history = ReportHistory(
                user_id=current_user.id,
                report_type='grupal',
                filename=os.path.basename(ruta_final),
                file_path=ruta_final,  # Usar ruta final en directorio de reportes
                file_size=file_size
            )
            db.session.add(report_history)
            db.session.commit()
            logger.info(f"Reporte grupal guardado en historial: {os.path.basename(ruta_final)}")
        except Exception as e:
            logger.error(f"Error guardando reporte grupal en historial: {str(e)}")
            db.session.rollback()
        
        # Devolver el archivo generado
        return send_file(
            ruta_final,  # Usar ruta final en directorio de reportes
            as_attachment=True,
            download_name=os.path.basename(ruta_final),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error en procesamiento de archivos grupales: {str(e)}")
        return Response(f'Error procesando archivos: {str(e)}', status=500)
